# -*- coding: utf-8 -*-
# ============================================================================
# 字段草案-v3.xlsx 的【渲染器】—— 数据权威源在 field-source.yaml（实现方案 D1）
#
# ⚠️ 本脚本自 P1 起不再内嵌字段数据：改字段请改 docs/standard/field-source.yaml，
# 然后运行：
#     python3 docs/standard/build_field_tables.py [可选:输出路径]
#     python3 docs/standard/check_field_source.py   # 校验 xlsx 与 YAML 一致（CI 强制）
# 依赖：openpyxl、pyyaml（pip install --break-system-packages openpyxl pyyaml）
#
# 四个 sheet：字段草案（实验记录 77 字段/9 模块）、一等实体字段表（MaterialLot/
# 装置Setup/表征仪器）、v2→v3变更说明、待明确清单。
# 全流程背景与进度见 docs/standard/STATUS.md；阶段边界见 docs/v2-implementation-plan.md。
# ============================================================================
import os
import sys

import openpyxl
import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "field-source.yaml")
OUT = sys.argv[1] if len(sys.argv) > 1 else os.path.join(BASE, "字段草案-v3.xlsx")

with open(SRC, encoding="utf-8") as fh:
    DOC = yaml.safe_load(fh)


def sections_to_rows(sections):
    """YAML sections -> 渲染行：("SEC", 标题) 或 10 列字段行（列序与表头一致）。"""
    rows = []
    for sec in sections:
        rows.append(("SEC", sec["title"]))
        for f in sec["fields"]:
            rows.append([
                f["module"], f["label"], f["meaning"], f["input"], f["options"],
                f["unit"], f["requirement"]["raw"], f["example"], f["source"], f["note"],
            ])
    return rows


HEADERS = DOC["experiment_record"]["headers"]
ROWS = sections_to_rows(DOC["experiment_record"]["sections"])
ENT_HEADERS = DOC["entities"]["headers"]
ENTITY_ROWS = sections_to_rows(DOC["entities"]["sections"])
CHG = DOC["changelog"]["rows"]
TBD = DOC["open_items"]["rows"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "字段草案"

# styles
hdr_fill = PatternFill("solid", fgColor="1F4E79")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
sec_fill = PatternFill("solid", fgColor="D6E4F0")
sec_font = Font(bold=True, color="1F4E79", size=11)
star_font = Font(color="C00000")
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
widths = [10, 16, 30, 12, 34, 10, 14, 26, 10, 34]


# 必填级别底色（前缀判定，与 req_font 同族）。「必填/选填」现行橙底为特例，须先于
# 通用「必填」前缀判断。其余「必填…」粉底、「条件必填…」橙底、「推荐…」黄底、「定义项」蓝底。
def req_fill(req):
    if req == "必填/选填":
        return PatternFill("solid", fgColor="FCEFD6")
    if req.startswith("必填"):
        return PatternFill("solid", fgColor="FCE4E4")
    if req.startswith("条件必填"):
        return PatternFill("solid", fgColor="FCEFD6")
    if req.startswith("推荐"):
        return PatternFill("solid", fgColor="FFF7DA")
    if req.startswith("定义项"):
        return PatternFill("solid", fgColor="E4ECF7")
    return None


# v3.4 导师提醒(B93)：必填/选填要有明显标识 → 必填级别加粗着色字体 + 表头下图例行
def req_font(req):
    if req.startswith("必填"):
        return Font(bold=True, color="C00000")
    if req.startswith("条件必填"):
        return Font(bold=True, color="B26B00")
    if req.startswith("推荐"):
        return Font(color="806000")
    return None


def add_legend(sheet):
    sheet.cell(2, 1, "图例▶").font = Font(bold=True, color="1F4E79")
    items = [("必填", "FCE4E4", "C00000"), ("条件必填(×××时)", "FCEFD6", "B26B00"),
             ("推荐", "FFF7DA", "806000"), ("选填(无底色)", "FFFFFF", "666666"),
             ("定义项(全局固定)", "E4ECF7", "1F4E79")]
    for i, (label, fg, fc) in enumerate(items, start=2):
        cell = sheet.cell(2, i, label)
        cell.fill = PatternFill("solid", fgColor=fg)
        cell.font = Font(bold=True, color=fc)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(2, 8, "表单UI另用红星*标必填（待实现，见待明确清单）").font = Font(color="666666", size=9)
    for c in range(1, 11):
        sheet.cell(2, c).border = border
    sheet.row_dimensions[2].height = 18


def render_field_sheet(sheet, headers, rows, *, start_row, freeze):
    """字段表 sheet 渲染（字段草案 / 一等实体字段表共用）：表头 + 章节行 + 字段行 + 列宽 + 冻结。"""
    for c, h in enumerate(headers, 1):
        cell = sheet.cell(1, c, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    r = start_row
    for item in rows:
        if isinstance(item, tuple):  # section
            sheet.cell(r, 1, item[1])
            sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
            cc = sheet.cell(r, 1)
            cc.fill = sec_fill
            cc.font = sec_font
            cc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            for c in range(1, 11):
                sheet.cell(r, c).border = border
            sheet.row_dimensions[r].height = 20
        else:
            for c, val in enumerate(item, 1):
                cell = sheet.cell(r, c, val)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            req = item[6]
            fill = req_fill(req)
            if fill:
                sheet.cell(r, 7).fill = fill
            rf = req_font(req)
            if rf:
                sheet.cell(r, 7).font = rf
            sheet.cell(r, 7).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            # star note red
            if str(item[9]).startswith("★"):
                sheet.cell(r, 10).font = star_font
        r += 1
    for c, w in enumerate(widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    sheet.freeze_panes = freeze


render_field_sheet(ws, HEADERS, ROWS, start_row=3, freeze="A3")
add_legend(ws)

# ---- 一等实体字段表 sheet ----
wsE = wb.create_sheet("一等实体字段表")
render_field_sheet(wsE, ENT_HEADERS, ENTITY_ROWS, start_row=2, freeze="A2")

# ---- 变更说明 sheet ----
ws2 = wb.create_sheet("v2→v3变更说明")
chg_hdr = DOC["changelog"]["headers"]
for c, h in enumerate(chg_hdr, 1):
    cell = ws2.cell(1, c, h)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
for i, row in enumerate(CHG, 2):
    for c, val in enumerate(row, 1):
        cell = ws2.cell(i, c, val)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
for c, w in enumerate([6, 16, 72, 14], 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
ws2.freeze_panes = "A2"

# ---- 待明确清单 sheet ----
ws3 = wb.create_sheet("待明确清单")
tbd_hdr = DOC["open_items"]["headers"]
for c, h in enumerate(tbd_hdr, 1):
    cell = ws3.cell(1, c, h)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
for i, row in enumerate(TBD, 2):
    for c, val in enumerate(row, 1):
        cell = ws3.cell(i, c, val)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
for c, w in enumerate([6, 64, 26], 1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
ws3.freeze_panes = "A2"

wb.save(OUT)
# count fields
nf = sum(1 for x in ROWS if not isinstance(x, tuple))
nreq = sum(1 for x in ROWS if not isinstance(x, tuple) and x[6] == "必填")
print("saved:", OUT)
print("字段行数:", nf, " 硬必填(必填)行:", nreq, " 章节:", sum(1 for x in ROWS if isinstance(x, tuple)))
