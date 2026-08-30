# -*- coding: utf-8 -*-
# ============================================================================
# field-source.yaml 一致性校验（实现方案 D1 红线的机器执行者，CI 强制）
#
# 校验三件事：
#   1. 提交的 字段草案-v3.xlsx 与「由 field-source.yaml 重新渲染」逐格一致（防手改 xlsx / 防漂移）
#   2. YAML 结构约束：必填级别词表封闭、条件必填必须带条件表达式、pending 字段必须有说明
#   3. 计数断言：字段数 / R0 数（防误删）
# 用法：uv run --project backend python docs/standard/check_field_source.py
# ============================================================================
import json
import os
import re
import subprocess
import sys
import tempfile

import openpyxl
import yaml

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "field-source.yaml")
COMMITTED = os.path.join(BASE, "字段草案-v3.xlsx")
RENDERER = os.path.join(BASE, "build_field_tables.py")

KNOWN_LEVELS = {
    "required",
    "recommended",
    "optional",
    "definition",
    "none",
    "mixed",
    "conditional_required",
    "conditional_recommended",
}
EXPECTED_FIELDS = 119
EXPECTED_ENTITY_FIELDS = 58
EXPECTED_R0 = 27

errors: list[str] = []


def err(msg: str) -> None:
    errors.append(msg)


# ---- 1. YAML 结构约束 ----
with open(SRC, encoding="utf-8") as fh:
    doc = yaml.safe_load(fh)


def iter_fields(part):
    for sec in doc[part]["sections"]:
        for f in sec["fields"]:
            yield f


n_fields = sum(1 for _ in iter_fields("experiment_record"))
n_entity = sum(1 for _ in iter_fields("entities"))
n_r0 = sum(1 for f in iter_fields("experiment_record") if f.get("r0"))

if n_fields != EXPECTED_FIELDS:
    err(
        f"实验记录字段数 {n_fields} ≠ 预期 {EXPECTED_FIELDS}（若为有意增删，请同步更新本脚本预期值并在 changelog 记录）"
    )
if n_entity != EXPECTED_ENTITY_FIELDS:
    err(f"一等实体字段数 {n_entity} ≠ 预期 {EXPECTED_ENTITY_FIELDS}（同上）")
if n_r0 != EXPECTED_R0:
    err(
        f"R0 标记数 {n_r0} ≠ 预期 {EXPECTED_R0}（R0 集合改动须导师/组会确认，见实现方案 §5）"
    )

for profile_key, profile in (doc.get("characterization_profiles") or {}).items():
    fields = {item["key"] for item in profile.get("condition_fields") or []}
    required = set(profile.get("required_condition_keys") or [])
    optional = set(profile.get("optional_condition_keys") or [])
    if required & optional or required | optional != fields:
        err(
            f"characterization_profiles.{profile_key} 的必填/选填测量条件未完整且互斥地覆盖 condition_fields"
        )

KEY_RE = re.compile(r"^[a-z][a-zA-Z0-9_]*$")  # 单位后缀允许大写（_C 等，沿 v1 风格）
modules_map = doc.get("modules", {})
entity_keys = doc.get("entity_keys", {})
stage_types = doc.get("stage_types", {})
group_names = set((stage_types.get("groups") or {}).keys())
ui_defaults = doc.get("field_ui_defaults") or {}
for key in (
    "input_placeholder",
    "input_placeholder_en",
    "select_placeholder",
    "select_placeholder_en",
):
    if not str(ui_defaults.get(key) or "").strip():
        err(f"field_ui_defaults.{key} 缺失或为空")
# 条件驱动字段原则上必须有 options；确需自由值驱动时在此显式列出。
CONDITION_OPTIONS_WHITELIST: set[str] = set()
# field_devices 跨实体驱动 process_steps.field_params，空值不走生成校验器；
# v2_experiment_service._validate_external_field_requirement 与前端均将 missing 判为不必填，实体服务另有“无”独占校验。
NE_DRIVER_OPTIONAL_ALLOWLIST = {"field_devices"}

seen_keys: dict[str, set] = {}
all_fields = [*iter_fields("experiment_record"), *iter_fields("entities")]


def field_scope(field):
    return modules_map.get(field["module"]) or entity_keys.get(field["module"])


def resolve_condition_field(raw_field):
    if not isinstance(raw_field, str) or "." not in raw_field:
        return None
    module, label = raw_field.split(".", 1)
    scope = modules_map.get(module) or entity_keys.get(module)
    if not scope:
        return None
    return next(
        (
            candidate
            for candidate in all_fields
            if field_scope(candidate) == scope and candidate["label"] == label
        ),
        None,
    )


for part, scope_of in (
    ("experiment_record", lambda f: modules_map.get(f["module"])),
    ("entities", lambda f: entity_keys.get(f["module"])),
):
    for f in iter_fields(part):
        where = f"{f['module']}/{f['label']}"
        req = f.get("requirement") or {}
        level = req.get("level")
        if level not in KNOWN_LEVELS:
            err(f"{where}: 未知必填级别 level={level!r}")
        if level in ("conditional_required", "conditional_recommended") and not req.get(
            "condition"
        ):
            err(f"{where}: 条件级别缺少 condition 表达式")
        condition = req.get("condition")
        otherwise = req.get("otherwise")
        if otherwise is not None and (
            otherwise != "optional" or level != "conditional_required"
        ):
            err(
                f"{where}: requirement.otherwise 仅支持 conditional_required + optional"
            )
        if condition and not {"field", "op", "value"} <= set(condition):
            err(f"{where}: condition 缺少 field/op/value")
        elif condition:
            driver = resolve_condition_field(condition["field"])
            if not driver:
                err(f"{where}: condition.field 无法解析: {condition['field']!r}")
            if condition["op"] not in {"eq", "ne", "in"}:
                err(f"{where}: condition.op 不支持: {condition['op']!r}")
            if condition["op"] == "in" and not isinstance(condition["value"], list):
                err(f"{where}: condition.op='in' 时 value 必须为 list")
            if driver:
                # ne 对空值的前后端语义可能分叉，除具名跨实体例外外，驱动字段必须必填。
                if (
                    condition["op"] == "ne"
                    and driver["key"] not in NE_DRIVER_OPTIONAL_ALLOWLIST
                    and driver["requirement"]["level"] != "required"
                ):
                    err(
                        f"{where}: condition.op='ne' 的驱动字段 {condition['field']!r} 必须为 required"
                    )
                raw_options = str(driver.get("options") or "").strip()
                if raw_options in {"", "—"}:
                    if condition["field"] not in CONDITION_OPTIONS_WHITELIST:
                        err(
                            f"{where}: condition 驱动字段 {condition['field']!r} 无 options 且未列入白名单"
                        )
                else:
                    separator = "·" if "·" in raw_options else "/"
                    options = {
                        doc.get("option_codes", {}).get(item.strip(), item.strip())
                        for item in raw_options.split(separator)
                        if item.strip()
                    }
                    raw_values = (
                        condition["value"]
                        if isinstance(condition["value"], list)
                        else [condition["value"]]
                    )
                    values = [
                        doc.get("option_codes", {}).get(value, value)
                        for value in raw_values
                    ]
                    unknown = [value for value in values if value not in options]
                    if unknown:
                        err(
                            f"{where}: condition 值 {unknown!r} 不在驱动字段 options 词表内"
                        )
        if "下拉" in str(f.get("input") or "") and str(
            f.get("options") or ""
        ).strip() in {"", "—"}:
            err(f"{where}: 下拉字段 options 必须非空且不能为 '—'")
        if f.get("status") == "pending-alignment" and not f.get("pending"):
            err(f"{where}: pending-alignment 缺少 pending 说明")
        validation = f.get("validation")
        if validation is not None:
            allowed_validation_keys = {
                "type",
                "ge",
                "gt",
                "le",
                "lt",
                "require_value",
                "require_value_options",
                "require_option",
                "item_required",
                "finite_value",
                "option_ranges",
                "sum",
                "tolerance",
                "unique_by",
            }
            unknown_validation = set(validation) - allowed_validation_keys
            if unknown_validation:
                err(f"{where}: validation 含未知约束 {sorted(unknown_validation)}")
            if validation.get("type") not in {None, "integer"}:
                err(f"{where}: validation.type 仅支持 integer")
            for flag in ("require_value", "require_option", "finite_value"):
                if flag in validation and not isinstance(validation[flag], bool):
                    err(f"{where}: validation.{flag} 必须为布尔值")
            require_value_options = validation.get("require_value_options")
            if require_value_options is not None and (
                not isinstance(require_value_options, list)
                or not require_value_options
                or not all(
                    isinstance(item, str) and item for item in require_value_options
                )
            ):
                err(f"{where}: validation.require_value_options 必须为非空字符串数组")
            for bound in ("ge", "gt", "le", "lt"):
                if bound in validation and (
                    isinstance(validation[bound], bool)
                    or not isinstance(validation[bound], (int, float))
                ):
                    err(f"{where}: validation.{bound} 必须为数值")
            for list_key in ("item_required", "unique_by"):
                items = validation.get(list_key)
                if items is not None and (
                    not isinstance(items, list)
                    or not items
                    or not all(isinstance(item, str) and item for item in items)
                ):
                    err(f"{where}: validation.{list_key} 必须为非空字符串数组")
            for numeric_key in ("sum", "tolerance"):
                value = validation.get(numeric_key)
                if value is not None and (
                    isinstance(value, bool) or not isinstance(value, (int, float))
                ):
                    err(f"{where}: validation.{numeric_key} 必须为数值")
            if (
                "ge" in validation
                and "le" in validation
                and validation["ge"] > validation["le"]
            ):
                err(f"{where}: validation.ge 不得大于 validation.le")
            option_ranges = validation.get("option_ranges")
            if option_ranges is not None:
                if not isinstance(option_ranges, dict) or not option_ranges:
                    err(f"{where}: validation.option_ranges 必须为非空对象")
                else:
                    raw_options = str(f.get("options") or "")
                    declared = {
                        doc.get("option_codes", {}).get(option.strip(), option.strip())
                        for option in raw_options.split("/")
                        if option.strip()
                    }
                    unknown_options = set(option_ranges) - declared
                    if unknown_options:
                        err(
                            f"{where}: validation.option_ranges 含未声明选项 "
                            f"{sorted(unknown_options)}"
                        )
                    unknown_required_options = (
                        set(require_value_options or []) - declared
                    )
                    if unknown_required_options:
                        err(
                            f"{where}: validation.require_value_options 含未声明选项 "
                            f"{sorted(unknown_required_options)}"
                        )
                    for option, bounds in option_ranges.items():
                        if not isinstance(bounds, dict) or not bounds:
                            err(
                                f"{where}: validation.option_ranges.{option} 必须为非空边界对象"
                            )
                            continue
                        unknown_bounds = set(bounds) - {"ge", "gt", "le", "lt"}
                        if unknown_bounds:
                            err(
                                f"{where}: validation.option_ranges.{option} "
                                f"含未知边界 {sorted(unknown_bounds)}"
                            )
                        for bound, value in bounds.items():
                            if isinstance(value, bool) or not isinstance(
                                value, (int, float)
                            ):
                                err(
                                    f"{where}: validation.option_ranges.{option}.{bound} 必须为数值"
                                )
        # D10: 机器字段键——必有、合法、模块/实体内唯一
        scope = scope_of(f)
        if scope is None:
            err(f"{where}: 模块 {f['module']!r} 未在 modules/entity_keys 映射中登记")
        key = f.get("key")
        if not key or not KEY_RE.match(str(key)):
            err(f"{where}: key 缺失或不合法: {key!r}")
        # D12: 字段层双语——英文名全量必填
        if not str(f.get("label_en") or "").strip():
            err(f"{where}: 缺少 label_en（国际化 D12 要求全量双语）")
        if bool(str(f.get("placeholder") or "").strip()) != bool(
            str(f.get("placeholder_en") or "").strip()
        ):
            err(f"{where}: placeholder / placeholder_en 必须成对填写")
        if bool(str(f.get("help") or "").strip()) != bool(
            str(f.get("help_en") or "").strip()
        ):
            err(f"{where}: help / help_en 必须成对填写")
        if scope:
            if key in seen_keys.setdefault(scope, set()):
                err(f"{where}: key {key!r} 在 {scope} 内重复")
            seen_keys[scope].add(key)
        # D11: §5 字段必须有参数组
        if part == "experiment_record" and scope == "process_steps":
            if f.get("group") not in group_names:
                err(
                    f"{where}: process_steps 字段缺少合法 group（现值 {f.get('group')!r}）"
                )

# D11: stage_types 自洽——shows ⊆ 组名；required_extra ⊆ §5 字段键
ps_keys = {
    f["key"]
    for f in iter_fields("experiment_record")
    if modules_map.get(f["module"]) == "process_steps"
}
for t in stage_types.get("types", []):
    bad = set(t.get("shows", [])) - group_names
    if bad:
        err(f"stage_types[{t.get('name')}]: 未知参数组 {sorted(bad)}")
    bad = set(t.get("required_extra", [])) - ps_keys
    if bad:
        err(
            f"stage_types[{t.get('name')}]: required_extra 引用不存在的字段键 {sorted(bad)}"
        )
if not stage_types.get("types"):
    err("缺少 stage_types.types（§5 动态表单权威映射，D11）")

# 科学合同：表征方法、属性单位和气体机器码必须自洽。
properties = doc.get("characterization_properties") or {}
property_units = (doc.get("scientific_contract") or {}).get("property_units") or {}
if set(properties) != set(property_units):
    err(
        "characterization_properties 与 scientific_contract.property_units "
        f"键不一致：{sorted(set(properties) ^ set(property_units))}"
    )
for property_code, definition in properties.items():
    where = f"characterization_properties.{property_code}"
    value_type = definition.get("value_type")
    if value_type not in {"numeric", "text", "structured"}:
        err(f"{where}: value_type 必须为 numeric / text / structured")
    validation = definition.get("validation")
    if not isinstance(validation, dict):
        err(f"{where}: validation 必须为对象")
        continue
    allowed = (
        {"ge", "gt", "le", "lt"}
        if value_type == "numeric"
        else {"min_length", "max_length"}
        if value_type == "text"
        else set()
    )
    unknown = set(validation) - allowed
    if unknown:
        err(f"{where}: validation 含不适用于 {value_type} 的约束 {sorted(unknown)}")
    for key, value in validation.items():
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            err(f"{where}: validation.{key} 必须为数值")
    if (
        "ge" in validation
        and "le" in validation
        and validation["ge"] > validation["le"]
    ):
        err(f"{where}: validation.ge 不得大于 validation.le")
    if (
        "min_length" in validation
        and "max_length" in validation
        and validation["min_length"] > validation["max_length"]
    ):
        err(f"{where}: validation.min_length 不得大于 validation.max_length")
if "layer_count" in properties:
    err("layer_count 只能作为 MaterialAssertion，不能同时作为 PropertyValue")
known_assertions = {
    "growth_presence",
    "phase_identity",
    "composition",
    "polytype",
    "stacking_order",
    "orientation_relationship",
    "layer_count",
}
component_keys = {
    "range": [{"min", "max"}, {"start", "end"}],
    "size": [{"x", "y"}, {"width", "height"}],
    "resolution": [{"width", "height"}],
}
for profile_code, profile in (doc.get("characterization_profiles") or {}).items():
    condition_fields = profile.get("condition_fields") or []
    condition_codes = [item.get("key") for item in condition_fields]
    if len(condition_codes) != len(set(condition_codes)):
        err(f"characterization_profiles.{profile_code}: 条件 key 重复")
    allowed_properties = set(profile.get("allowed_property_codes") or [])
    unknown_properties = allowed_properties - set(properties)
    if unknown_properties:
        err(
            f"characterization_profiles.{profile_code}: 未知属性 {sorted(unknown_properties)}"
        )
    if not set(profile.get("default_property_codes") or []) <= allowed_properties:
        err(f"characterization_profiles.{profile_code}: 默认属性不属于允许属性")
    unknown_assertions = (
        set(profile.get("allowed_assertion_types") or []) - known_assertions
    )
    if unknown_assertions:
        err(
            f"characterization_profiles.{profile_code}: 未知材料结论 {sorted(unknown_assertions)}"
        )
    for field in condition_fields:
        value_type = field.get("value_type")
        components = field.get("components") or []
        validation = field.get("validation") or {}
        if not isinstance(validation, dict):
            err(
                f"characterization_profiles.{profile_code}.{field.get('key')}: "
                "validation 必须为对象"
            )
            validation = {}
        allowed_validation = (
            {"min_length", "max_length"}
            if value_type == "text"
            else {"ge", "gt", "le", "lt"}
        )
        unknown_validation = set(validation) - allowed_validation
        if unknown_validation:
            err(
                f"characterization_profiles.{profile_code}.{field.get('key')}: "
                f"validation 含不适用于 {value_type} 的约束 "
                f"{sorted(unknown_validation)}"
            )
        for key, value in validation.items():
            if isinstance(value, bool) or not isinstance(value, (int, float)):
                err(
                    f"characterization_profiles.{profile_code}.{field.get('key')}: "
                    f"validation.{key} 必须为数值"
                )
        if (
            "min_length" in validation
            and "max_length" in validation
            and validation["min_length"] > validation["max_length"]
        ):
            err(
                f"characterization_profiles.{profile_code}.{field.get('key')}: "
                "validation.min_length 不得大于 validation.max_length"
            )
        if value_type in component_keys:
            keys = {item.get("key") for item in components}
            if keys not in component_keys[value_type]:
                err(
                    f"characterization_profiles.{profile_code}.{field.get('key')}: "
                    f"{value_type} 分量 {sorted(keys)} 不合法"
                )
        elif components:
            err(
                f"characterization_profiles.{profile_code}.{field.get('key')}: "
                "标量条件不能声明 components"
            )
if not doc.get("characterization_profiles"):
    err("缺少 characterization_profiles")

gas_aliases: dict[str, str] = {}
for code, definition in (doc.get("gas_species") or {}).items():
    for alias in definition.get("aliases") or []:
        normalized = str(alias).strip().casefold()
        previous = gas_aliases.get(normalized)
        if previous and previous != code:
            err(f"gas_species 别名 {alias!r} 同时映射到 {previous} 与 {code}")
        gas_aliases[normalized] = code
if not doc.get("gas_species"):
    err("缺少 gas_species 受控词表")

# ---- 2. xlsx 逐格一致 ----
if not os.path.exists(COMMITTED):
    err(f"缺少提交的 xlsx: {COMMITTED}")
else:
    with tempfile.TemporaryDirectory() as td:
        regen = os.path.join(td, "regen.xlsx")
        proc = subprocess.run(
            [sys.executable, RENDERER, regen], capture_output=True, text=True
        )
        if proc.returncode != 0:
            err(f"渲染器运行失败：{proc.stderr.strip()[:500]}")
        else:
            wa = openpyxl.load_workbook(COMMITTED)
            wb = openpyxl.load_workbook(regen)
            for sheet_name, fields, first_data_row in (
                ("字段草案", list(iter_fields("experiment_record")), 3),
                ("一等实体字段表", list(iter_fields("entities")), 2),
            ):
                sheet = wa[sheet_name]
                headers = [cell.value for cell in sheet[1]]
                if not headers or headers[-1] != "机器约束":
                    err(f"[{sheet_name}] 缺少末列『机器约束』")
                    continue
                machine_column = len(headers)
                actual = [
                    sheet.cell(row, machine_column).value or ""
                    for row in range(first_data_row, sheet.max_row + 1)
                    if sheet.cell(row, 2).value is not None
                ]
                expected = [
                    (
                        json.dumps(
                            field["validation"],
                            ensure_ascii=False,
                            sort_keys=True,
                            separators=(",", ":"),
                        )
                        if field.get("validation")
                        else ""
                    )
                    for field in fields
                ]
                if actual != expected:
                    err(f"[{sheet_name}] 机器约束列与 YAML validation 不一致")
            if wa.sheetnames != wb.sheetnames:
                err(f"sheet 列表不一致：{wa.sheetnames} vs {wb.sheetnames}")
            for name in wa.sheetnames:
                if len(errors) > 40:
                    break
                if name not in wb.sheetnames:
                    continue
                sa, sb = wa[name], wb[name]
                max_r = max(sa.max_row, sb.max_row)
                max_c = max(sa.max_column, sb.max_column)
                for r in range(1, max_r + 1):
                    for c in range(1, max_c + 1):
                        va, vb = sa.cell(r, c).value, sb.cell(r, c).value
                        if (va or "") != (vb or ""):
                            err(f"[{name}] r{r}c{c}: 提交版={va!r} ≠ 重新生成={vb!r}")
                            if len(errors) > 40:
                                err("……diff 过多，截断")
                                break
                    if len(errors) > 40:
                        break

if errors:
    print("❌ field-source 校验失败：")
    for e in errors:
        print("  -", e)
    print(
        "修复方式：改 field-source.yaml → uv run --project backend python "
        "docs/standard/build_field_tables.py → 重新提交 xlsx。"
    )
    sys.exit(1)

print(
    f"✅ field-source 校验通过：字段 {n_fields} · 实体字段 {n_entity} · R0 {n_r0} · xlsx 与 YAML 渲染逐格一致"
)
