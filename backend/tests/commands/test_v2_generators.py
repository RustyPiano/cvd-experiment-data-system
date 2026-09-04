import importlib.util
import sys
from copy import deepcopy
from datetime import date
from pathlib import Path
from types import ModuleType
from typing import get_args
from uuid import uuid4

import openpyxl
import pytest
from jsonschema import Draft202012Validator

from app.commands.export_v2_schema import export_v2_schema
from app.commands.generate_v2_models import (
    generate_material_phase_catalog,
    generate_v2_models,
    render_v2_models,
)
from app.main import app
from app.schemas.generated.v2_module_payload import (
    V2_MODULE_PAYLOAD_MODELS,
    ActualFieldPayload,
    InstrumentVersionPayload,
    MaterialLotReferencePayload,
    MaterialLotVersionPayload,
    SetupVersionPayload,
    validate_v2_module_payload,
)
from app.schemas.scientific import (
    MeasurementConditions,
    SampleRegion,
)
from app.services.v2_field_source import load_field_source, missing

REPO_ROOT = Path(__file__).resolve().parents[3]
GENERATED_SCHEMA_DIR = REPO_ROOT / "docs" / "standard" / "generated"
GENERATED_MODEL = REPO_ROOT / "backend" / "app" / "schemas" / "generated" / "v2_module_payload.py"
GENERATED_PHASE_CATALOG = REPO_ROOT / "backend" / "app" / "generated" / "material_phase_catalog.py"
LOT_REF = {
    "entity_id": "7d9e7787-e5ef-4f34-818f-454a10263a3b",
    "version": 2,
}


def _sensors(count: int = 2) -> list[dict]:
    return [
        {
            "sensor_type": "thermocouple",
            "zone_index": zone_index,
            "nominal_accuracy_C": 1.0,
        }
        for zone_index in range(1, count + 1)
    ]


def _equipment(**changes) -> dict:
    payload = {
        "setup_ref": "setup-1@v2",
        "setup_origin": "commercial",
        "manufacturer_brand": "Vendor",
        "zone_count": 2,
        "orientation": "horizontal",
        "tube_material_shape": {"material": "quartz", "shape": "round"},
        "tube_outer_diameter_wall_mm": {
            "outer_diameter_mm": 50.0,
            "wall_thickness_mm": 2.0,
        },
        "temperature_sensors": _sensors(),
        "tube_usage_history": {
            "reset_count": 0,
            "use_number_since_reset": 1,
        },
    }
    payload.update(changes)
    return payload


def _setup(**changes) -> dict:
    payload = {
        "setup_name": "Two-zone furnace",
        "setup_origin": "commercial",
        "manufacturer_brand": "Vendor",
        "setup_code": "SETUP-1",
        "zone_count": 2,
        "temperature_sensors": _sensors(),
        "orientation": "horizontal",
        "tube_material_shape": {"material": "quartz", "shape": "round"},
        "tube_outer_diameter_wall_mm": {
            "outer_diameter_mm": 50.0,
            "wall_thickness_mm": 2.0,
        },
        "field_devices": ["none"],
    }
    payload.update(changes)
    return payload


def _substrate(**changes) -> dict:
    item = {
        "material": "sapphire_al2o3",
        "lot_ref": LOT_REF,
        "piece_label": "S1",
        "chemical_formula": "Al2O3",
        "size_placement": {
            "length_mm": 10.0,
            "width_mm": 10.0,
            "placement": "face_up",
        },
        "zone_thermocouple_distance_mm": {
            "zone_index": 1,
            "distance_mm": 0.0,
        },
    }
    item.update(changes)
    return item


@pytest.fixture
def rendered_models(tmp_path: Path) -> ModuleType:
    target = tmp_path / "rendered_v2_module_payload.py"
    target.write_text(render_v2_models(load_field_source()), encoding="utf-8")
    spec = importlib.util.spec_from_file_location("_rendered_v2_module_payload", target)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def test_generate_v2_models_matches_committed_file(tmp_path: Path) -> None:
    generated = generate_v2_models(output_path=tmp_path / "v2_module_payload.py")

    assert Path(generated).read_bytes() == GENERATED_MODEL.read_bytes()
    assert set(V2_MODULE_PAYLOAD_MODELS) == {"equipment", "substrates"}


def test_generate_material_phase_catalog_matches_committed_file(tmp_path: Path) -> None:
    generated = generate_material_phase_catalog(output_path=tmp_path / "material_phase_catalog.py")

    assert Path(generated).read_bytes() == GENERATED_PHASE_CATALOG.read_bytes()


def test_export_v2_schema_matches_committed_files(tmp_path: Path) -> None:
    paths = export_v2_schema(output_dir=tmp_path)

    expected = {
        "json_schema": "cvd-2d-process-v2.schema.json",
        "field_dictionary_json": "cvd-2d-field-dictionary-v2.json",
    }
    for key, filename in expected.items():
        assert Path(paths[key]).read_bytes() == (GENERATED_SCHEMA_DIR / filename).read_bytes()


def test_measurement_json_schema_enforces_property_and_profile_contract() -> None:
    exported = export_v2_schema(output_dir=None)
    schema = exported["json_schema_doc"]["result_models"]["measurement_bundle"]
    validator = Draft202012Validator(schema)
    payload = {
        "measurement": {
            "sample_id": "00000000-0000-0000-0000-000000000001",
            "method_profile": "optical_microscopy",
            "measured_at": "2026-08-30T12:00:00+00:00",
            "sample_region": {
                "geometry_type": "whole_sample",
                "label": "whole sample",
                "coordinate_system": "sample_local",
            },
            "typed_conditions": {},
        },
        "properties": [
            {
                "property_code": "coverage_percent",
                "numeric_value": 80,
                "unit": "%",
                "quality_flag": "valid",
            }
        ],
    }
    assert not list(validator.iter_errors(payload))

    wrong_representation = deepcopy(payload)
    wrong_representation["properties"][0] = {
        "property_code": "coverage_percent",
        "text_value": "negative",
    }
    assert list(validator.iter_errors(wrong_representation))

    out_of_range = deepcopy(payload)
    out_of_range["properties"][0]["numeric_value"] = 101
    assert list(validator.iter_errors(out_of_range))

    numeric_below_limit = deepcopy(payload)
    numeric_below_limit["properties"][0]["quality_flag"] = "below_detection_limit"
    assert not list(validator.iter_errors(numeric_below_limit))

    text_property = deepcopy(payload)
    text_property["properties"][0] = {
        "property_code": "observation_note",
        "text_value": "no visible feature",
        "quality_flag": "valid",
    }
    assert not list(validator.iter_errors(text_property))
    text_property["properties"][0]["quality_flag"] = "below_detection_limit"
    assert list(validator.iter_errors(text_property))

    wrong_profile = deepcopy(payload)
    wrong_profile["measurement"]["method_profile"] = "other"
    assert list(validator.iter_errors(wrong_profile))

    missing_evidence = deepcopy(payload)
    missing_evidence["properties"] = []
    assert list(validator.iter_errors(missing_evidence))

    other = {
        "measurement": {
            "sample_id": "00000000-0000-0000-0000-000000000001",
            "method_profile": "other",
            "measured_at": "2026-08-30T12:00:00+00:00",
            "sample_region": {
                "geometry_type": "point",
                "label": "measurement point",
                "coordinate_system": "sample_local",
            },
            "typed_conditions": {"method_description": "custom spectroscopy"},
            "raw_file_ids": ["00000000-0000-0000-0000-000000000002"],
        }
    }
    assert not list(validator.iter_errors(other))
    other["measurement"]["typed_conditions"]["method_description"] = "   "
    assert list(validator.iter_errors(other))
    other["measurement"]["typed_conditions"]["method_description"] = "x" * 1001
    assert list(validator.iter_errors(other))

    assert exported["json_schema_doc"]["characterization_properties"]
    assert exported["field_dictionary_doc"]["characterization_profiles"]


def test_measurement_json_schema_and_openapi_reject_runtime_invalid_shapes() -> None:
    standalone = export_v2_schema(output_dir=None)["json_schema_doc"]["result_models"][
        "measurement_bundle"
    ]
    openapi = app.openapi()
    validators = [
        Draft202012Validator(standalone),
        Draft202012Validator(
            {
                "$ref": "#/components/schemas/MeasurementBundleCreate",
                "components": openapi["components"],
            }
        ),
    ]
    payload = {
        "measurement": {
            "sample_id": "00000000-0000-0000-0000-000000000001",
            "method_profile": "optical_microscopy",
            "measured_at": "2026-08-30T12:00:00+00:00",
            "sample_region": {
                "geometry_type": "whole_sample",
                "label": "whole sample",
                "coordinate_system": "sample_local",
            },
            "typed_conditions": {},
        },
        "properties": [
            {
                "property_code": "coverage_percent",
                "numeric_value": 80,
                "unit": "%",
            }
        ],
    }
    invalid_payloads: list[dict] = []

    required_condition_is_null = deepcopy(payload)
    required_condition_is_null["measurement"].update(
        {
            "method_profile": "other",
            "sample_region": {
                "geometry_type": "point",
                "label": "point",
                "coordinate_system": "sample_local",
            },
            "typed_conditions": {"method_description": None},
            "raw_file_ids": ["00000000-0000-0000-0000-000000000002"],
        }
    )
    required_condition_is_null["properties"] = []
    invalid_payloads.append(required_condition_is_null)

    line_with_height = deepcopy(payload)
    line_with_height["measurement"].update(
        {
            "method_profile": "Raman",
            "instrument_id": "00000000-0000-0000-0000-000000000003",
            "instrument_version": 1,
            "sample_region": {
                "geometry_type": "line",
                "label": "line",
                "coordinate_system": "sample_local",
                "width": 2,
                "height": 1,
                "unit": "um",
            },
            "typed_conditions": {"laser_wavelength_nm": 532},
            "raw_file_ids": ["00000000-0000-0000-0000-000000000002"],
        }
    )
    line_with_height["properties"] = []
    invalid_payloads.append(line_with_height)

    whole_sample_with_coordinates = deepcopy(payload)
    whole_sample_with_coordinates["measurement"]["sample_region"].update(
        {"x": 1, "y": 2, "unit": "um"}
    )
    invalid_payloads.append(whole_sample_with_coordinates)

    malformed_assertion = deepcopy(payload)
    malformed_assertion["assertions"] = [
        {
            "assertion_type": "phase_identity",
            "value": {"phase": " ", "extra": "not allowed"},
        }
    ]
    invalid_payloads.append(malformed_assertion)

    duplicate_analysis_input = deepcopy(payload)
    duplicate_analysis_input["analyses"] = [
        {
            "software_name": "analysis",
            "software_version": "1",
            "started_at": "2026-08-30T12:00:00+00:00",
            "input_file_ids": [
                "00000000-0000-0000-0000-000000000002",
                "00000000-0000-0000-0000-000000000002",
            ],
        }
    ]
    invalid_payloads.append(duplicate_analysis_input)

    for validator in validators:
        assert not list(validator.iter_errors(payload))
        for invalid_payload in invalid_payloads:
            assert list(validator.iter_errors(invalid_payload))


def test_field_dictionary_conditions_use_machine_codes() -> None:
    dictionary = export_v2_schema(output_dir=None)["field_dictionary_doc"]
    values = [
        item
        for field in dictionary["fields"]
        if field["condition"] is not None
        for item in (
            field["condition"]["value"]
            if isinstance(field["condition"]["value"], list)
            else [field["condition"]["value"]]
        )
    ]

    assert all(not isinstance(value, str) or value.isascii() for value in values)
    assert set(values) >= {
        "chemical",
        "staged_cooling",
        "gas_cylinder",
        "gas_line",
        "open_lid_cooling",
        "sio2_si",
        "substrate",
    }


def test_characterization_field_source_matches_runtime_discriminators() -> None:
    profiles = load_field_source()["characterization_profiles"].values()

    profile_condition_fields = {
        field["key"] for profile in profiles for field in profile["condition_fields"]
    }
    assert profile_condition_fields == set(MeasurementConditions.model_fields) - {"power_setting"}
    assert {
        assertion for profile in profiles for assertion in profile["allowed_assertion_types"]
    } == set()
    assert {region for profile in profiles for region in profile["allowed_region_types"]} == set(
        get_args(SampleRegion.model_fields["geometry_type"].annotation)
    )


def test_published_field_dictionary_paths_and_machine_types_match_schema() -> None:
    exported = export_v2_schema(output_dir=None)
    schema = exported["json_schema_doc"]
    fields = exported["field_dictionary_doc"]["fields"]
    paths: set[str] = set()

    for field in fields:
        path = field["schema_path"]
        assert path is not None, f"{field['module_key']}.{field['key']} has no released schema path"
        assert path not in paths, f"duplicate released schema path: {path}"
        paths.add(path)
        prefix = "entities" if field["source_part"] == "entity" else "modules"
        module_schema = schema[prefix][field["module_key"]]
        relative_path = path.removeprefix(f"$.{prefix}.{field['module_key']}.")
        nodes = _schema_nodes_at_path(module_schema, relative_path)
        machine_types = {
            item_type
            for node in nodes
            for item_type in _schema_node_types(node, module_schema)
            if item_type != "null"
        }
        assert field["machine_type"] == " | ".join(sorted(machine_types))

    by_key = {(field["module_key"], field["key"]): field for field in fields}
    assert by_key[("process_steps", "channels")]["schema_path"] == (
        "$.modules.process_steps.properties.channels"
    )
    assert by_key[("substrates", "material")]["schema_path"] == (
        "$.modules.substrates.properties.items.items.properties.material"
    )
    assert by_key[("precursors", "unit")]["schema_path"] == (
        "$.modules.precursors.properties.items.items.properties.ingredients.items.properties.unit"
    )
    assert by_key[("target_product", "target_layer_count")]["schema_path"] == (
        "$.modules.target_product.properties.material_regions.items.properties.target_layer_count"
    )


def _schema_nodes_at_path(root: dict, path: str) -> list[dict]:
    nodes = [root]
    for token in path.split("."):
        nodes = [
            value
            for node in nodes
            for variant in _schema_node_variants(node, root)
            if isinstance(value := variant.get(token), dict)
        ]
        assert nodes, f"schema path does not exist: {path}"
    return nodes


def _schema_node_variants(node: dict, root: dict) -> list[dict]:
    ref = node.get("$ref")
    if isinstance(ref, str) and ref.startswith("#/"):
        resolved = root
        for token in ref[2:].split("/"):
            resolved = resolved[token.replace("~1", "/").replace("~0", "~")]
        node = resolved
    variants = [
        *(node.get("oneOf") or []),
        *(node.get("anyOf") or []),
        *(node.get("allOf") or []),
    ]
    return [item for variant in variants for item in _schema_node_variants(variant, root)] or [node]


def _schema_node_types(node: dict, root: dict) -> set[str]:
    return {
        item_type
        for variant in _schema_node_variants(node, root)
        for item_type in (
            [variant["type"]] if isinstance(variant.get("type"), str) else variant.get("type", [])
        )
        if isinstance(item_type, str)
    }


def test_standard_schema_exports_current_scientific_and_result_models() -> None:
    schema = export_v2_schema(output_dir=None)["json_schema_doc"]

    assert set(schema["modules"]) == {
        "basic_info",
        "equipment",
        "precursors",
        "process_events",
        "process_steps",
        "substrates",
        "target_product",
    }
    assert "created_by_user_id" in schema["modules"]["basic_info"]["properties"]
    assert "material_regions" in schema["modules"]["target_product"]["properties"]
    assert "channels" in schema["modules"]["process_steps"]["properties"]
    assert set(schema["result_models"]) == {
        "measurement_bundle",
        "transformation",
        "dataset_query",
    }
    assert schema["version"] == "v4.0-alpha.40"
    assert schema["status"] == "INTERNAL_VALIDATION"
    tilt_schema = schema["modules"]["substrates"]["$defs"]["SubstrateSizePlacementPayload"]
    assert tilt_schema["properties"]["tilt_angle_deg"]["anyOf"][0]["not"] == {"const": 0}
    assert "pvd" not in schema["modules"]
    for model in [*schema["modules"].values(), *schema["result_models"].values()]:
        Draft202012Validator.check_schema(model)


def test_field_dictionary_and_xlsx_expose_machine_validation_contract() -> None:
    dictionary = export_v2_schema(output_dir=None)["field_dictionary_doc"]
    assert all(
        field["module_key"] not in {"measured_products", "characterization", "pvd"}
        for field in dictionary["fields"]
    )

    workbook = openpyxl.load_workbook(REPO_ROOT / "docs" / "standard" / "字段草案-v3.xlsx")
    sheet = workbook["字段草案"]
    headers = [cell.value for cell in sheet[1]]
    assert headers[-1] == "机器约束"
    machine_column = len(headers)
    amount_row = next(
        row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 2).value == "使用量"
    )
    assert '"gt":0' in sheet.cell(amount_row, machine_column).value


def test_process_timeline_schema_validates_the_current_top_level_contract() -> None:
    schema = export_v2_schema(output_dir=None)["json_schema_doc"]["modules"]["process_steps"]
    validator = Draft202012Validator(schema)
    payload = {
        "segments": [],
        "channels": [
            {
                "channel_key": "channel_11111111_1111_4111_8111_111111111111",
                "channel_type": "temperature",
                "source_type": "setpoint",
                "subject_type": "temperature_zone",
                "subject_ref": "zone_1",
                "subject_instance_ref": "zone_1_controller",
                "zone_index": 1,
                "unit": "℃",
                "data_kind": "scalar",
                "scalar_value": 750,
            }
        ],
        "pressure_regime": "atmospheric",
        "cooling_method": "furnace_cooling",
    }

    assert validator.is_valid(payload)
    assert not validator.is_valid({**payload, "channels": []})


def test_material_lot_enforces_current_category_specific_fields() -> None:
    chemical = MaterialLotVersionPayload.model_validate(
        {
            "lot_category": "chemical",
            "substance_name": "Molybdenum trioxide",
            "chemical_formula": " Mo O₃ ",
            "batch_number": "MO-1",
        }
    )
    assert chemical.chemical_formula == "MoO3"

    substrate = MaterialLotVersionPayload.model_validate(
        {
            "lot_category": "substrate",
            "substance_name": "Sapphire",
            "chemical_formula": "Al2O3",
            "batch_number_availability": "batch_number_reported",
            "batch_number": "S-1",
            "substrate_material": "sapphire_al2o3",
        }
    )
    assert substrate.substrate_material == "sapphire_al2o3"

    no_batch_substrate = MaterialLotVersionPayload.model_validate(
        {
            "lot_category": "substrate",
            "substance_name": "Sapphire",
            "chemical_formula": "Al2O3",
            "batch_number_availability": "batch_number_not_provided",
            "production_date": "2026-08",
            "substrate_material": "sapphire_al2o3",
        }
    )
    assert no_batch_substrate.batch_number is None
    assert no_batch_substrate.production_date == "2026-08"

    with pytest.raises(ValueError, match="production_date"):
        MaterialLotVersionPayload.model_validate(
            {
                "lot_category": "substrate",
                "substance_name": "Sapphire",
                "chemical_formula": "Al2O3",
                "batch_number_availability": "batch_number_not_provided",
                "substrate_material": "sapphire_al2o3",
            }
        )

    with pytest.raises(ValueError, match="batch_number must match"):
        MaterialLotVersionPayload.model_validate(
            {
                "lot_category": "substrate",
                "substance_name": "Sapphire",
                "chemical_formula": "Al2O3",
                "batch_number_availability": "batch_number_not_provided",
                "batch_number": "INVENTED",
                "substrate_material": "sapphire_al2o3",
            }
        )

    with pytest.raises(ValueError, match="chemical_formula"):
        MaterialLotVersionPayload.model_validate(
            {
                "lot_category": "chemical",
                "substance_name": "Unknown",
                "batch_number": "C-1",
            }
        )
    with pytest.raises(ValueError, match="substrate_material"):
        MaterialLotVersionPayload.model_validate(
            {
                "lot_category": "substrate",
                "substance_name": "Unknown substrate",
                "chemical_formula": "Si",
                "batch_number_availability": "batch_number_reported",
                "batch_number": "S-2",
            }
        )


def test_gas_cylinder_requires_a_valid_fixed_composition() -> None:
    pure = MaterialLotVersionPayload.model_validate(
        {
            "lot_category": "gas_cylinder",
            "substance_name": "Carbon dioxide",
            "batch_number": "CO2-1",
            "gas_components": [{"species": "CO2", "volume_percent": 100}],
        }
    )
    assert pure.gas_components[0].species == "CO2"

    mixed = MaterialLotVersionPayload.model_validate(
        {
            "lot_category": "gas_cylinder",
            "substance_name": "5% H2 / Ar",
            "batch_number": "MIX-1",
            "gas_components": [
                {"species": "H2", "volume_percent": 5},
                {"species": "Ar", "volume_percent": 95},
            ],
        }
    )
    assert sum(item.volume_percent for item in mixed.gas_components) == 100

    for components, message in (
        ([{"species": "Ar", "volume_percent": 90}], "sum to 100"),
        (
            [
                {"species": "Ar", "volume_percent": 50},
                {"species": "Ar", "volume_percent": 50},
            ],
            "unique",
        ),
    ):
        with pytest.raises(ValueError, match=message):
            MaterialLotVersionPayload.model_validate(
                {
                    "lot_category": "gas_cylinder",
                    "substance_name": "Invalid gas",
                    "batch_number": "BAD-1",
                    "gas_components": components,
                }
            )

    with pytest.raises(ValueError, match="not applicable"):
        MaterialLotVersionPayload.model_validate(
            {
                "lot_category": "gas_cylinder",
                "substance_name": "Argon",
                "chemical_formula": "Ar",
                "batch_number": "AR-1",
                "gas_components": [{"species": "Ar", "volume_percent": 100}],
            }
        )


def test_setup_and_instrument_entity_contracts() -> None:
    setup = SetupVersionPayload.model_validate(_setup())
    assert setup.orientation == "horizontal"

    named_field_setup = SetupVersionPayload.model_validate(
        _setup(field_devices=["other"], field_device_other_name="magnetic field")
    )
    assert named_field_setup.field_device_other_name == "magnetic field"
    with pytest.raises(ValueError, match="field_device_other_name"):
        SetupVersionPayload.model_validate(_setup(field_devices=["other"]))

    actual_other_field = ActualFieldPayload.model_validate(
        {
            "field_type": "other",
            "start_min": 5,
            "end_min": 25,
            "parameters": [{"name": "field_strength", "value": 0.5, "unit": "T"}],
        }
    )
    assert actual_other_field.field_type == "other"

    instrument = InstrumentVersionPayload.model_validate(
        {
            "instrument_code": "RAMAN-1",
            "name_type": "Raman spectrometer",
            "capabilities": [{"method": "raman"}],
            "last_calibration": date(2026, 8, 31),
        }
    )
    assert instrument.instrument_code == "RAMAN-1"
    reference_id = uuid4()
    assert (
        MaterialLotReferencePayload.model_validate(
            {"entity_id": reference_id, "version": 1}
        ).entity_id
        == reference_id
    )

    for invalid in (float("nan"), float("inf"), object()):
        with pytest.raises(ValueError, match="finite JSON"):
            InstrumentVersionPayload.model_validate(
                {
                    "instrument_code": "RAMAN-BAD",
                    "name_type": "Raman spectrometer",
                    "capabilities": [{"value": invalid}],
                }
            )

    with pytest.raises(ValueError, match="cover each zone"):
        SetupVersionPayload.model_validate(_setup(temperature_sensors=_sensors(1)))
    with pytest.raises(ValueError, match="not applicable"):
        SetupVersionPayload.model_validate(
            _setup(
                setup_origin="lab_built",
                design_build_organization="Lab",
                internal_model="LAB-1",
                manufacturer_brand="Vendor",
            )
        )


def test_equipment_and_setup_share_geometry_and_zone_validation(
    rendered_models: ModuleType,
) -> None:
    for model, payload in (
        (rendered_models.EquipmentPayload, _equipment()),
        (rendered_models.SetupVersionPayload, _setup()),
    ):
        model.model_validate(payload)
        with pytest.raises(ValueError, match="cover each zone"):
            model.model_validate({**payload, "temperature_sensors": _sensors(1)})
        with pytest.raises(ValueError, match="cross-section shape"):
            model.model_validate(
                {
                    **payload,
                    "tube_outer_diameter_wall_mm": {
                        "outer_width_mm": 50,
                        "outer_height_mm": 30,
                        "wall_thickness_mm": 2,
                    },
                }
            )


def test_equipment_snapshot_round_trips_current_fields() -> None:
    validated = validate_v2_module_payload("equipment", _equipment())

    assert validated["orientation"] == "horizontal"
    assert validated["temperature_sensors"][1]["zone_index"] == 2
    assert validated["tube_usage_history"] == {
        "reset_count": 0,
        "use_number_since_reset": 1,
    }


def test_substrate_module_validates_current_geometry_and_material_conditions() -> None:
    validated = validate_v2_module_payload("substrates", {"items": [_substrate()]})
    assert validated["items"][0]["chemical_formula"] == "Al2O3"

    sio2 = _substrate(
        material="sio2_si",
        chemical_formula="SiO2",
        oxide_thickness_nm=285,
    )
    validate_v2_module_payload("substrates", {"items": [sio2]})
    with pytest.raises(ValueError, match="oxide_thickness_nm"):
        validate_v2_module_payload(
            "substrates",
            {"items": [_substrate(material="sio2_si", chemical_formula="SiO2")]},
        )
    with pytest.raises(ValueError, match="not applicable"):
        validate_v2_module_payload(
            "substrates",
            {"items": [_substrate(oxide_thickness_nm=285)]},
        )


def test_substrate_tilt_and_pretreatment_use_structured_contracts() -> None:
    tilted = _substrate(
        size_placement={
            "length_mm": 10,
            "width_mm": 10,
            "placement": "tilted",
            "tilt_angle_deg": -15,
            "tilt_azimuth_deg": 180,
        },
        pretreatment_steps=[
            {
                "type": "plasma_treatment",
                "parameters": {
                    "power_W": 50,
                    "gas_species": "O2",
                    "duration_min": 2,
                    "pressure_Pa": 20,
                },
            }
        ],
    )
    saved = validate_v2_module_payload("substrates", {"items": [tilted]})["items"][0]
    assert saved["size_placement"]["tilt_angle_deg"] == -15
    assert saved["size_placement"]["tilt_azimuth_deg"] == 180
    assert saved["pretreatment_steps"][0]["type"] == "plasma_treatment"

    cleaned = validate_v2_module_payload(
        "substrates",
        {
            "items": [
                _substrate(
                    pretreatment_steps=[
                        {
                            "type": "solvent_cleaning",
                            "parameters": {
                                "solvent": "acetone",
                                "cleaning_method": "ultrasonic",
                                "duration_min": 10,
                            },
                        },
                        {
                            "type": "uv_ozone_treatment",
                            "parameters": {"duration_min": 15},
                        },
                    ]
                )
            ]
        },
    )
    assert [step["type"] for step in cleaned["items"][0]["pretreatment_steps"]] == [
        "solvent_cleaning",
        "uv_ozone_treatment",
    ]

    with pytest.raises(ValueError, match="duration_min"):
        validate_v2_module_payload(
            "substrates",
            {
                "items": [
                    _substrate(
                        pretreatment_steps=[
                            {
                                "type": "solvent_cleaning",
                                "parameters": {
                                    "solvent": "acetone",
                                    "cleaning_method": "ultrasonic",
                                },
                            }
                        ]
                    )
                ]
            },
        )

    with pytest.raises(ValueError, match="tilt_angle_deg"):
        validate_v2_module_payload(
            "substrates",
            {
                "items": [
                    _substrate(
                        size_placement={
                            "length_mm": 10,
                            "width_mm": 10,
                            "placement": "tilted",
                        }
                    )
                ]
            },
        )

    with pytest.raises(ValueError, match="non-zero"):
        validate_v2_module_payload(
            "substrates",
            {
                "items": [
                    _substrate(
                        size_placement={
                            "length_mm": 10,
                            "width_mm": 10,
                            "placement": "tilted",
                            "tilt_angle_deg": 0,
                            "tilt_azimuth_deg": 180,
                        }
                    )
                ]
            },
        )

    with pytest.raises(ValueError, match="length_mm"):
        validate_v2_module_payload(
            "substrates",
            {
                "items": [
                    _substrate(
                        size_placement={
                            "length_mm": 5,
                            "width_mm": 10,
                            "placement": "face_up",
                        }
                    )
                ]
            },
        )

    related = validate_v2_module_payload(
        "substrates",
        {
            "items": [_substrate(), _substrate(piece_label="S2")],
            "placement_relations": [{"piece_a_label": "S1", "piece_b_label": "S2", "gap_mm": 0}],
        },
    )
    assert related["placement_relations"][0]["gap_mm"] == 0
    with pytest.raises(ValueError, match="two different pieces"):
        validate_v2_module_payload(
            "substrates",
            {
                "items": [_substrate()],
                "placement_relations": [{"piece_a_label": "S1", "piece_b_label": "S1"}],
            },
        )


def test_required_value_detection_keeps_false_and_zero() -> None:
    assert missing("   ")
    assert missing({"value": None, "option": None})
    assert not missing(False)
    assert not missing(0)


def test_generator_fails_closed_for_unknown_complex_input_type() -> None:
    doc = deepcopy(load_field_source())
    field = next(
        item
        for section in doc["experiment_record"]["sections"]
        for item in section["fields"]
        if item["key"] == "manufacturer_brand"
    )
    field["input"] = "未实现的复杂对象"

    with pytest.raises(ValueError, match="Unsupported input type"):
        render_v2_models(doc)
