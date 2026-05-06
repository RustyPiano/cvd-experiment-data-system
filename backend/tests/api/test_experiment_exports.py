from io import BytesIO
from uuid import UUID

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.models.module_payload import ExperimentModuleKey, ExperimentModulePayload

client = TestClient(app)


def login(email: str, password: str = "Password123!") -> str:
    response = client.post(
        "/api/v1/auth/login",
        json={"email": email, "password": password},
    )
    assert response.status_code == 200
    return response.json()["access_token"]


def auth_headers(email: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {login(email)}"}


def create_experiment(email: str, *, objective: str = "Export flow") -> str:
    response = client.post(
        "/api/v1/experiments",
        json={
            "experiment_type": "cvd_2zone",
            "material_system": "MoS2",
            "experiment_date": "2026-04-23",
            "objective": objective,
        },
        headers=auth_headers(email),
    )
    assert response.status_code == 201
    return response.json()["id"]


def populate_required_modules(experiment_id: str, email: str) -> None:
    precursors_response = client.put(
        f"/api/v1/experiments/{experiment_id}/modules/precursors",
        json={"payload_json": {"items": [{"species": "MoO3", "method": "powder"}]}},
        headers=auth_headers(email),
    )
    assert precursors_response.status_code == 200

    furnace_response = client.put(
        f"/api/v1/experiments/{experiment_id}/modules/furnace_program",
        json={
            "payload_json": {
                "furnace_info": {"zones_count": 1, "initial_temperatures_C": {"zone_1": 25}},
                "precursors": [],
                "steps": [
                    {
                        "step_index": 1,
                        "step_name": "升温",
                        "duration_min": 30,
                        "is_hold": False,
                        "temperatures_C": {"zone_1": 750},
                        "note": "",
                    },
                ],
            }
        },
        headers=auth_headers(email),
    )
    assert furnace_response.status_code == 200

    gas_response = client.put(
        f"/api/v1/experiments/{experiment_id}/modules/gas_program",
        json={
            "payload_json": {
                "segments": [
                    {
                        "stage": "growth",
                        "start_min": 0,
                        "end_min": 45,
                        "gas": "Ar",
                        "components": [{"name": "Ar", "fraction": 1, "flow_sccm": 80}],
                        "flow_sccm": 80,
                    }
                ]
            }
        },
        headers=auth_headers(email),
    )
    assert gas_response.status_code == 200


def create_product_sample(experiment_id: str, email: str) -> str:
    response = client.post(
        f"/api/v1/experiments/{experiment_id}/samples",
        json={"role": "product", "storage_location": "drawer-1"},
        headers=auth_headers(email),
    )
    assert response.status_code == 201
    return response.json()["id"]


def sync_top_substrate_sample(experiment_id: str, email: str) -> str:
    substrates_response = client.put(
        f"/api/v1/experiments/{experiment_id}/modules/substrates",
        json={
            "payload_json": {
                "items": [
                    {
                        "role": "top",
                        "type": "SiO2/Si",
                        "brand": "MTI",
                        "size_mm": "10x10",
                        "treatment_method": "acetone",
                        "position_mm": 12.5,
                    }
                ]
            }
        },
        headers=auth_headers(email),
    )
    assert substrates_response.status_code == 200

    samples_response = client.get(
        f"/api/v1/samples?experiment_id={experiment_id}&role=top",
        headers=auth_headers(email),
    )
    assert samples_response.status_code == 200
    samples = samples_response.json()["items"]
    assert len(samples) == 1
    return samples[0]["id"]


def sync_top_and_bottom_substrates(experiment_id: str, email: str) -> dict[str, str]:
    substrates_response = client.put(
        f"/api/v1/experiments/{experiment_id}/modules/substrates",
        json={
            "payload_json": {
                "items": [
                    {
                        "role": "top",
                        "type": "SiO2/Si",
                        "brand": "MTI",
                        "size_mm": "10x10",
                        "treatment_method": "plasma_cleaning",
                        "position_mm": 12.5,
                    },
                    {
                        "role": "bottom",
                        "type": "Quartz",
                        "brand": "Kejing",
                        "size_mm": "10x10",
                        "treatment_method": "annealing",
                        "position_mm": -12.5,
                    },
                ]
            }
        },
        headers=auth_headers(email),
    )
    assert substrates_response.status_code == 200

    samples_response = client.get(
        f"/api/v1/samples?experiment_id={experiment_id}",
        headers=auth_headers(email),
    )
    assert samples_response.status_code == 200
    return {item["role"]: item["id"] for item in samples_response.json()["items"]}


def test_export_includes_related_sample_and_file_audit_events(active_user) -> None:
    experiment_id = create_experiment(active_user.email)
    sample_id = sync_top_substrate_sample(experiment_id, active_user.email)

    sample_update_response = client.patch(
        f"/api/v1/samples/{sample_id}",
        json={"storage_location": "box-a1"},
        headers=auth_headers(active_user.email),
    )
    assert sample_update_response.status_code == 200

    file_response = client.post(
        f"/api/v1/experiments/{experiment_id}/files",
        headers=auth_headers(active_user.email),
        data={
            "sample_id": sample_id,
            "method": "Raman",
            "file_category": "raw",
            "note": "linked spectrum",
        },
        files={"file": ("linked-raman.txt", b"peak=404", "text/plain")},
    )
    assert file_response.status_code == 201
    file_id = file_response.json()["id"]

    export_response = client.get(
        f"/api/v1/experiments/{experiment_id}/export",
        headers=auth_headers(active_user.email),
    )

    assert export_response.status_code == 200
    audit_events = export_response.json()["audit_events"]
    entity_types = {event["entity_type"] for event in audit_events}
    assert {"experiment_run", "sample", "file_asset"} <= entity_types
    assert any(event["entity_type"] != "experiment_run" for event in audit_events)

    sample_events = [
        event
        for event in audit_events
        if event["entity_type"] == "sample" and event["entity_id"] == sample_id
    ]
    assert {"create", "update"} <= {event["action"] for event in sample_events}

    file_events = [
        event
        for event in audit_events
        if event["entity_type"] == "file_asset" and event["entity_id"] == file_id
    ]
    assert {"create"} <= {event["action"] for event in file_events}


def test_export_retains_soft_deleted_substrate_sample_provenance(active_user) -> None:
    experiment_id = create_experiment(active_user.email, objective="Soft deleted sample export")
    sample_ids = sync_top_and_bottom_substrates(experiment_id, active_user.email)
    bottom_sample_id = sample_ids["bottom"]

    remove_bottom_response = client.put(
        f"/api/v1/experiments/{experiment_id}/modules/substrates",
        json={
            "payload_json": {
                "items": [
                    {
                        "role": "top",
                        "type": "SiO2/Si",
                        "brand": "MTI",
                        "size_mm": "10x10",
                        "treatment_method": "plasma_cleaning",
                        "position_mm": 12.5,
                    }
                ]
            }
        },
        headers=auth_headers(active_user.email),
    )
    assert remove_bottom_response.status_code == 200

    export_response = client.get(
        f"/api/v1/experiments/{experiment_id}/export",
        headers=auth_headers(active_user.email),
    )
    analysis_response = client.get(
        f"/api/v1/experiments/{experiment_id}/export/analysis",
        headers=auth_headers(active_user.email),
    )

    assert export_response.status_code == 200
    assert analysis_response.status_code == 200
    export_body = export_response.json()
    bottom_sample = next(
        sample for sample in export_body["samples"] if sample["id"] == bottom_sample_id
    )
    assert bottom_sample["is_deleted"] is True
    assert bottom_sample["deleted_at"] is not None
    assert bottom_sample["deleted_by_id"] is not None

    bottom_events = [
        event
        for event in export_body["audit_events"]
        if event["entity_type"] == "sample" and event["entity_id"] == bottom_sample_id
    ]
    assert "soft_delete" in {event["action"] for event in bottom_events}

    analysis_body = analysis_response.json()
    bottom_row = next(
        row for row in analysis_body["sample_rows"] if row["sample_id"] == bottom_sample_id
    )
    assert bottom_row["is_deleted"] is True
    assert bottom_row["deleted_at"] is not None


def test_export_experiment_aggregates_modules_samples_files_and_audit(active_user) -> None:
    experiment_id = create_experiment(active_user.email)

    substrates_response = client.put(
        f"/api/v1/experiments/{experiment_id}/modules/substrates",
        json={
            "payload_json": {
                "items": [
                    {
                        "role": "top",
                        "type": "SiO2/Si",
                        "brand": "MTI",
                        "size_mm": "10x10",
                        "treatment_method": "acetone",
                        "position_mm": 12.5,
                    }
                ]
            }
        },
        headers=auth_headers(active_user.email),
    )
    assert substrates_response.status_code == 200
    populate_required_modules(experiment_id, active_user.email)
    sample_id = create_product_sample(experiment_id, active_user.email)

    keep_file_response = client.post(
        f"/api/v1/experiments/{experiment_id}/files",
        headers=auth_headers(active_user.email),
        data={
            "sample_id": sample_id,
            "method": "Raman",
            "file_category": "raw",
            "note": "primary spectrum",
        },
        files={"file": ("raman.txt", b"peak=404", "text/plain")},
    )
    assert keep_file_response.status_code == 201

    deleted_file_response = client.post(
        f"/api/v1/experiments/{experiment_id}/files",
        headers=auth_headers(active_user.email),
        data={"method": "OM"},
        files={"file": ("draft-note.txt", b"remove-me", "text/plain")},
    )
    assert deleted_file_response.status_code == 201
    deleted_file_id = deleted_file_response.json()["id"]
    delete_response = client.delete(
        f"/api/v1/files/{deleted_file_id}",
        headers=auth_headers(active_user.email),
    )
    assert delete_response.status_code == 204

    export_response = client.get(
        f"/api/v1/experiments/{experiment_id}/export",
        headers=auth_headers(active_user.email),
    )

    assert export_response.status_code == 200
    body = export_response.json()
    assert body["export_version"] == "cvd_export_v1"
    assert body["experiment"]["id"] == experiment_id
    assert body["features"] == []
    assert body["provenance"] == {
        "derived_from_run_id": None,
        "derived_from_run_code": None,
    }
    assert body["counts"] == {
        "modules": 4,
        "samples": 2,
        "files": 1,
        "audit_events": 13,
    }
    assert {item["module_key"] for item in body["modules"]} == {
        "substrates",
        "precursors",
        "furnace_program",
        "gas_program",
    }
    assert {item["sample_code"] for item in body["samples"]} == {
        "S-2026-0001-TOP",
        "S-2026-0001-PRODUCT-A",
    }
    assert [item["original_name"] for item in body["files"]] == ["raman.txt"]
    assert body["files"][0]["sample_id"] == sample_id
    assert body["files"][0]["method"] == "Raman"
    assert "upload_file" in {item["action"] for item in body["audit_events"]}
    assert "delete_file" in {item["action"] for item in body["audit_events"]}
    deleted_file_events = [
        event
        for event in body["audit_events"]
        if event["entity_type"] == "file_asset" and event["entity_id"] == deleted_file_id
    ]
    assert {"create", "delete"} <= {event["action"] for event in deleted_file_events}


def test_export_analysis_returns_normalized_rows(active_user) -> None:
    experiment_id = create_experiment(active_user.email, objective="Analysis export")

    precursors_response = client.put(
        f"/api/v1/experiments/{experiment_id}/modules/precursors",
        json={
            "payload_json": {
                "items": [
                    {
                        "species": "MoO3",
                        "brand": "Sigma",
                        "concentration": 0.25,
                        "concentration_unit": "mol/L",
                        "method": "solution",
                        "melting_temperature_C": 795,
                        "spin_speed_rpm": 3000,
                        "pre_spin_speed_rpm": 500,
                        "preparation_time_min": 20,
                        "mass_mg": 12.5,
                        "batch_no": "MO-0424",
                    }
                ]
            }
        },
        headers=auth_headers(active_user.email),
    )
    assert precursors_response.status_code == 200

    substrates_response = client.put(
        f"/api/v1/experiments/{experiment_id}/modules/substrates",
        json={
            "payload_json": {
                "items": [
                    {
                        "role": "top",
                        "type": "SiO2/Si",
                        "brand": "MTI",
                        "size_mm": "10x10",
                        "treatment_method": "plasma",
                        "position_mm": 12.5,
                        "treatment_params": {
                            "temperature_C": 80,
                            "duration_min": 5,
                            "power_W": 30,
                            "gas": "O2",
                        },
                    }
                ]
            }
        },
        headers=auth_headers(active_user.email),
    )
    assert substrates_response.status_code == 200

    furnace_response = client.put(
        f"/api/v1/experiments/{experiment_id}/modules/furnace_program",
        json={
            "payload_json": {
                "furnace_info": {"zones_count": 1, "initial_temperatures_C": {"zone_1": 25}},
                "precursors": [
                    {"material": "", "position_cm": None, "mass_mg": None, "note": "center"}
                ],
                "steps": [
                    {
                        "step_index": 1,
                        "step_name": "升温",
                        "duration_min": 30,
                        "is_hold": False,
                        "temperatures_C": {"zone_1": 750},
                        "note": "",
                    },
                ],
            }
        },
        headers=auth_headers(active_user.email),
    )
    assert furnace_response.status_code == 200

    gas_response = client.put(
        f"/api/v1/experiments/{experiment_id}/modules/gas_program",
        json={
            "payload_json": {
                "pre_washing_gas": "Ar",
                "segments": [
                    {
                        "stage": "growth",
                        "start_min": 0,
                        "end_min": 45,
                        "gas": "Ar/H2",
                        "flow_sccm": 80,
                        "note": "stable",
                        "components": [
                            {"name": "Ar", "fraction": 0.9, "flow_sccm": 72},
                            {"name": "H2", "ratio_percent": 10, "flow_sccm": 8},
                        ],
                    }
                ],
            }
        },
        headers=auth_headers(active_user.email),
    )
    assert gas_response.status_code == 200

    characterization_response = client.put(
        f"/api/v1/experiments/{experiment_id}/modules/characterization",
        json={
            "payload_json": {
                "methods": [
                    {
                        "method": "Raman",
                        "enabled": True,
                        "excitation_nm": 532,
                        "result": "E2g/A1g visible",
                        "note": "room temperature",
                    }
                ]
            }
        },
        headers=auth_headers(active_user.email),
    )
    assert characterization_response.status_code == 200

    sample_id = create_product_sample(experiment_id, active_user.email)
    file_response = client.post(
        f"/api/v1/experiments/{experiment_id}/files",
        headers=auth_headers(active_user.email),
        data={
            "sample_id": sample_id,
            "method": "Raman",
            "file_category": "raw",
            "note": "analysis input",
        },
        files={"file": ("analysis-raman.txt", b"peak=404", "text/plain")},
    )
    assert file_response.status_code == 201

    export_response = client.get(
        f"/api/v1/experiments/{experiment_id}/export/analysis",
        headers=auth_headers(active_user.email),
    )

    assert export_response.status_code == 200
    body = export_response.json()
    assert body["export_version"] == "cvd_analysis_v1"
    assert body["experiment"]["experiment_id"] == experiment_id
    assert body["experiment"]["run_code"] == "CVD-2026-0001"
    assert body["experiment"]["material_system"] == "MoS2"

    expected_sections = {
        "experiment",
        "precursor_rows",
        "substrate_rows",
        "furnace_step_rows",
        "furnace_precursor_rows",
        "gas_program_rows",
        "gas_segment_rows",
        "gas_component_rows",
        "characterization_rows",
        "sample_rows",
        "file_rows",
    }
    assert expected_sections <= set(body)

    for section in expected_sections - {"experiment"}:
        assert body[section]
        assert {
            "experiment_id": experiment_id,
            "run_code": "CVD-2026-0001",
        }.items() <= body[section][0].items()

    assert body["precursor_rows"][0] == {
        "experiment_id": experiment_id,
        "run_code": "CVD-2026-0001",
        "precursor_index": 0,
        "species": "MoO3",
        "brand": "Sigma",
        "concentration": 0.25,
        "concentration_unit": "mol/L",
        "method": "solution",
        "melting_temperature_C": 795.0,
        "spin_speed_rpm": 3000.0,
        "pre_spin_speed_rpm": 500.0,
        "preparation_time_min": 20.0,
        "mass_mg": 12.5,
        "batch_no": "MO-0424",
    }
    assert body["substrate_rows"][0]["substrate_index"] == 0
    assert body["substrate_rows"][0]["treatment_params_gas"] == "O2"
    assert body["furnace_step_rows"][0]["step_index"] == 1
    assert body["furnace_step_rows"][0]["temperature_C"] == 750.0
    assert body["gas_program_rows"][0] == {
        "experiment_id": experiment_id,
        "run_code": "CVD-2026-0001",
        "gas_program_index": 0,
        "pre_washing_gas": "Ar",
    }
    assert body["gas_segment_rows"][0]["gas_segment_index"] == 0
    assert body["gas_segment_rows"][0]["pre_washing_gas"] == "Ar"
    assert {row["component_name"] for row in body["gas_component_rows"]} == {"Ar", "H2"}
    assert body["characterization_rows"][0]["characterization_index"] == 0
    assert body["characterization_rows"][0]["excitation_nm"] == 532.0
    assert {row["sample_code"] for row in body["sample_rows"]} == {
        "S-2026-0001-TOP",
        "S-2026-0001-PRODUCT-A",
    }
    assert body["file_rows"][0]["file_id"] == file_response.json()["id"]
    assert body["file_rows"][0]["sample_id"] == sample_id


def test_export_analysis_keeps_program_level_gas_and_flat_metadata(active_user) -> None:
    experiment_id = create_experiment(active_user.email, objective="Flat analysis export")

    gas_response = client.put(
        f"/api/v1/experiments/{experiment_id}/modules/gas_program",
        json={"payload_json": {"pre_washing_gas": "Ar", "segments": []}},
        headers=auth_headers(active_user.email),
    )
    assert gas_response.status_code == 200

    sample_response = client.post(
        f"/api/v1/experiments/{experiment_id}/samples",
        json={
            "role": "product",
            "storage_location": "drawer-2",
            "metadata_json": {"nested": {"grade": "A"}, "tags": ["analysis"]},
        },
        headers=auth_headers(active_user.email),
    )
    assert sample_response.status_code == 201
    sample_id = sample_response.json()["id"]

    file_response = client.post(
        f"/api/v1/experiments/{experiment_id}/files",
        headers=auth_headers(active_user.email),
        data={
            "sample_id": sample_id,
            "method": "OM",
            "file_category": "processed",
        },
        files={"file": ("flat-metadata.txt", b"same-bytes", "text/plain")},
    )
    assert file_response.status_code == 201

    export_response = client.get(
        f"/api/v1/experiments/{experiment_id}/export/analysis",
        headers=auth_headers(active_user.email),
    )

    assert export_response.status_code == 200
    body = export_response.json()
    assert body["gas_program_rows"] == [
        {
            "experiment_id": experiment_id,
            "run_code": "CVD-2026-0001",
            "gas_program_index": 0,
            "pre_washing_gas": "Ar",
        }
    ]
    assert body["gas_segment_rows"] == []
    assert body["sample_rows"][0]["metadata_json_text"] == (
        '{"nested":{"grade":"A"},"tags":["analysis"]}'
    )
    assert "metadata_json" not in body["sample_rows"][0]
    assert body["file_rows"][0]["metadata_json_text"] == "{}"
    assert "metadata_json" not in body["file_rows"][0]


def test_export_analysis_rejects_legacy_invalid_numeric_payload(active_user, db_session) -> None:
    experiment_id = create_experiment(active_user.email, objective="Legacy analysis export")
    populate_required_modules(experiment_id, active_user.email)

    db_session.add(
        ExperimentModulePayload(
            experiment_run_id=UUID(experiment_id),
            module_key=ExperimentModuleKey.CHARACTERIZATION.value,
            payload_json={
                "methods": [{"method": "Raman", "enabled": True, "excitation_nm": "bad"}]
            },
        )
    )
    db_session.commit()

    non_raising_client = TestClient(app, raise_server_exceptions=False)
    export_response = non_raising_client.get(
        f"/api/v1/experiments/{experiment_id}/export/analysis",
        headers=auth_headers(active_user.email),
    )

    assert export_response.status_code == 422
    detail = export_response.json()["detail"]
    assert any(
        error["module_key"] == "characterization"
        and error["field_path"] == "methods[0].excitation_nm"
        and "valid number" in error["message"]
        for error in detail
    )


def test_export_experiment_excel_returns_openable_workbook(active_user) -> None:
    experiment_id = create_experiment(active_user.email, objective="Excel export flow")
    populate_required_modules(experiment_id, active_user.email)

    file_response = client.post(
        f"/api/v1/experiments/{experiment_id}/files",
        headers=auth_headers(active_user.email),
        data={"method": "OM", "file_category": "processed"},
        files={"file": ("image.png", b"png-bytes", "image/png")},
    )
    assert file_response.status_code == 201

    export_response = client.get(
        f"/api/v1/experiments/{experiment_id}/export/excel",
        headers=auth_headers(active_user.email),
    )

    assert export_response.status_code == 200
    assert (
        export_response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(export_response.content))
    assert workbook.sheetnames == [
        "Basic Info",
        "Environment & Precheck",
        "Precursors",
        "Substrates",
        "Furnace Program",
        "Gas Program",
        "Characterization",
        "Files",
        "Audit",
    ]
    assert workbook["Basic Info"]["A1"].value == "Field"
    assert workbook["Basic Info"]["B2"].value == "CVD-2026-0001"
    assert workbook["Files"]["A2"].value == "image.png"
    assert workbook["Files"]["B2"].value == "OM"
    assert workbook["Files"]["C2"].value == "processed"
    assert [cell.value for cell in workbook["Audit"][1]] == [
        "created_at",
        "entity_type",
        "entity_id",
        "action",
        "actor_id",
        "reason",
    ]
    audit_rows = list(workbook["Audit"].iter_rows(min_row=2, values_only=True))
    assert any(row[1] == "file_asset" and row[3] == "create" for row in audit_rows)


def test_viewer_can_export_locked_experiment_but_not_other_users_draft(
    admin_user, viewer_user
) -> None:
    locked_experiment_id = create_experiment(admin_user.email, objective="Visible export")
    populate_required_modules(locked_experiment_id, admin_user.email)

    submit_response = client.post(
        f"/api/v1/experiments/{locked_experiment_id}/submit",
        headers=auth_headers(admin_user.email),
    )
    assert submit_response.status_code == 200

    lock_response = client.post(
        f"/api/v1/experiments/{locked_experiment_id}/lock",
        headers=auth_headers(admin_user.email),
    )
    assert lock_response.status_code == 200

    visible_response = client.get(
        f"/api/v1/experiments/{locked_experiment_id}/export",
        headers=auth_headers(viewer_user.email),
    )
    assert visible_response.status_code == 200
    assert visible_response.json()["experiment"]["status"] == "locked"

    draft_experiment_id = create_experiment(admin_user.email, objective="Hidden draft export")
    hidden_response = client.get(
        f"/api/v1/experiments/{draft_experiment_id}/export",
        headers=auth_headers(viewer_user.email),
    )

    assert hidden_response.status_code == 404
