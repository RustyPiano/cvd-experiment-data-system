from datetime import time
from typing import Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.imports import ParsedExperimentDraft
from app.services.imports.base import ImportProfile


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def _to_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _time_to_minutes(value: Any) -> float | None:
    """Best-effort convert a spreadsheet cell to minutes.

    The lab's machine stores gas timings as time-of-day cells (HH:MM:SS), which
    we fold into total minutes. Plain numbers are treated as already-in-minutes.
    """
    if value is None or value == "":
        return None
    if isinstance(value, time):
        return value.hour * 60 + value.minute + value.second / 60
    return _to_float(value)


class CvdProcessPackageProfile(ImportProfile):
    key = "cvd_process_package_v1"
    display_name = "自动化 CVD 工艺参数包"
    description = (
        "实验室自动化 CVD 机台导出的宽表（每行一条实验：A/B 前驱体、炉温台阶、四路气体）。"
    )

    GASES = ("Ar", "H2", "O2", "CO2")

    def parse(self, workbook: Workbook) -> tuple[list[ParsedExperimentDraft], list[str]]:
        worksheet: Worksheet = workbook.worksheets[0]
        header_index = self._build_header_index(worksheet)
        if "Order" not in header_index and "A" not in header_index:
            return [], ["未识别到工艺参数包表头（缺少 Order / A 等列），请确认文件格式。"]

        drafts: list[ParsedExperimentDraft] = []
        for row in range(2, worksheet.max_row + 1):

            def cell(name: str, _row: int = row) -> Any:
                col = header_index.get(name)
                if col is None:
                    return None
                return worksheet.cell(_row, col).value

            if self._row_is_empty(cell, header_index):
                continue
            drafts.append(self._parse_row(row, cell))

        if not drafts:
            return [], ["未在文件中找到任何数据行。"]
        return drafts, []

    def _build_header_index(self, worksheet: Worksheet) -> dict[str, int]:
        header_index: dict[str, int] = {}
        for col in range(1, worksheet.max_column + 1):
            header = worksheet.cell(1, col).value
            if header is None:
                continue
            header_index[str(header).strip()] = col
        return header_index

    def _row_is_empty(self, cell, header_index: dict[str, int]) -> bool:
        for name in ("Order", "A", "B", "Substrate"):
            if name in header_index and cell(name) not in (None, ""):
                return False
        return True

    def _parse_row(self, row: int, cell) -> ParsedExperimentDraft:
        warnings: list[str] = []

        precursors_payload = self._build_precursors(cell)
        substrates_payload, substrate_warnings = self._build_substrates(cell)
        furnace_payload, furnace_warnings = self._build_furnace(cell, precursors_payload)
        gas_payload, gas_warnings = self._build_gas(cell)
        warnings.extend(substrate_warnings)
        warnings.extend(furnace_warnings)
        warnings.extend(gas_warnings)

        order = _to_text(cell("Order"))
        run_level: dict[str, Any] = {
            "experiment_type": "cvd_2zone",
            "material_system": None,
            "objective": f"导入自工艺参数包 第 {order} 条" if order else "导入自工艺参数包",
        }

        module_payloads: dict[str, Any] = {
            "precursors": precursors_payload,
            "substrates": substrates_payload,
            "furnace_program": furnace_payload,
            "gas_program": gas_payload,
        }
        return ParsedExperimentDraft(
            source_row=row,
            run_level=run_level,
            module_payloads=module_payloads,
            warnings=warnings,
        )

    def _build_precursors(self, cell) -> dict[str, Any]:
        items: list[dict[str, Any]] = []
        for species_col, mass_col in (("A", "mass_A"), ("B", "mass_B")):
            species = _to_text(cell(species_col))
            mass = _to_float(cell(mass_col))
            if species is None and mass is None:
                continue
            items.append(
                {
                    "species": species,
                    "brand": None,
                    "method": None,
                    "mass_mg": mass,
                    "batch_no": None,
                }
            )
        return {"items": items}

    def _build_substrates(self, cell) -> tuple[dict[str, Any], list[str]]:
        substrate_type = _to_text(cell("Substrate"))
        if substrate_type is None:
            return {"items": []}, []
        warnings = ["基底角色默认为“下基底”，请在确认时核对。"]
        item = {
            "role": "bottom",
            "type": substrate_type,
            "brand": None,
            "size_mm": None,
            "treatment_method": None,
            "position_mm": None,
        }
        return {"items": [item]}, warnings

    def _build_furnace(
        self,
        cell,
        precursors_payload: dict[str, Any],
    ) -> tuple[dict[str, Any], list[str]]:
        warnings: list[str] = []
        zone_a_nodes = self._zone_nodes(
            cell,
            prefix="A",
            temperature_steps=6,
            end_time_key="A_end_time",
            warnings=warnings,
        )
        zone_b_nodes = self._zone_nodes(
            cell,
            prefix="B",
            temperature_steps=4,
            end_time_key="B_end_time",
            warnings=warnings,
        )

        zones: list[dict[str, Any]] = []
        if zone_a_nodes:
            zones.append({"zone_key": "zone_1", "temperature_program": zone_a_nodes, "note": ""})
        if zone_b_nodes:
            zones.append({"zone_key": "zone_2", "temperature_program": zone_b_nodes, "note": ""})

        initial_temperatures: dict[str, float] = {}
        if zone_a_nodes and zone_a_nodes[0].get("temperature_C") is not None:
            initial_temperatures["zone_1"] = zone_a_nodes[0]["temperature_C"]
        if zone_b_nodes and zone_b_nodes[0].get("temperature_C") is not None:
            initial_temperatures["zone_2"] = zone_b_nodes[0]["temperature_C"]

        placements: list[dict[str, Any]] = []
        precursor_count = len(precursors_payload.get("items", []))
        if precursor_count >= 1 and zone_a_nodes:
            placements.append(
                {"precursor_index": 0, "zone_key": "zone_1", "position_cm": None, "note": ""}
            )
        if precursor_count >= 2 and zone_b_nodes:
            placements.append(
                {"precursor_index": 1, "zone_key": "zone_2", "position_cm": None, "note": ""}
            )

        if zones:
            warnings.append(
                "炉温时间已按“每步为相邻温度间的升温时长”累加为绝对时间，请确认。"
            )

        payload = {
            "furnace_info": {
                "zones_count": 2,
                "model": None,
                "initial_temperatures_C": initial_temperatures or None,
            },
            "placements": placements,
            "zones": zones,
        }
        return payload, warnings

    def _zone_nodes(
        self,
        cell,
        *,
        prefix: str,
        temperature_steps: int,
        end_time_key: str,
        warnings: list[str],
    ) -> list[dict[str, Any]]:
        # Each step{i}_time is the ramp duration from temperature{i} to
        # temperature{i+1}, so absolute node times are the running sum of those
        # durations (node 1 sits at t=0). A final end-time adds a hold at the
        # last temperature.
        temperatures = [
            _to_float(cell(f"{prefix}_step{step}_temperature"))
            for step in range(1, temperature_steps + 1)
        ]
        ramp_durations = [
            _to_float(cell(f"{prefix}_step{step}_time"))
            for step in range(1, temperature_steps)
        ]

        nodes: list[dict[str, Any]] = []
        cumulative = 0.0
        for index, temperature in enumerate(temperatures):
            if index > 0:
                duration = ramp_durations[index - 1]
                if duration is None:
                    duration = 0.0
                elif duration < 0:
                    warnings.append(
                        f"{prefix} 温区第 {index} 步升温时间为负值（{duration}），已按 0 处理，请确认。"
                    )
                    duration = 0.0
                cumulative += duration
            if temperature is None:
                continue
            nodes.append(
                {
                    "node_index": len(nodes) + 1,
                    "time_min": cumulative,
                    "temperature_C": temperature,
                    "note": "",
                }
            )

        end_time = _to_float(cell(end_time_key))
        if end_time is not None and end_time > 0 and nodes:
            cumulative += end_time
            nodes.append(
                {
                    "node_index": len(nodes) + 1,
                    "time_min": cumulative,
                    "temperature_C": nodes[-1]["temperature_C"],
                    "note": "保温/结束",
                }
            )
        return nodes

    def _build_gas(self, cell) -> tuple[dict[str, Any], list[str]]:
        segments: list[dict[str, Any]] = []
        any_time_cells = False
        for gas in self.GASES:
            step1_flow = _to_float(cell(f"{gas}_step1_flow"))
            end_flow = _to_float(cell(f"{gas}_end_flow"))
            if not step1_flow and not end_flow:
                continue
            start_min = _time_to_minutes(cell(f"{gas}_step1_time"))
            end_min = _time_to_minutes(cell(f"{gas}_step3_time"))
            if cell(f"{gas}_step1_time") is not None or cell(f"{gas}_step3_time") is not None:
                any_time_cells = True
            segments.append(
                {
                    "stage": "growth",
                    "gas": gas,
                    "start_min": start_min,
                    "end_min": end_min,
                    "flow_sccm": step1_flow if step1_flow else end_flow,
                    "note": "",
                    "components": [{"name": gas, "gas": gas, "flow_sccm": step1_flow}],
                }
            )

        warnings: list[str] = []
        if any_time_cells:
            warnings.append(
                "气体时间单元格按 时:分:秒 折算为分钟，流量与时间请在确认时核对。"
            )
        return {"pre_washing_gas": None, "segments": segments}, warnings
