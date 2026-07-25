from __future__ import annotations

import csv
import math
from io import BytesIO, StringIO
from numbers import Real
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.models.file_asset import FileAsset
from app.services.file_storage_service import FileStorageService

TIMESERIES_METADATA_KEYS = (
    "columns",
    "numeric_columns",
    "numeric_column_pairs",
    "row_count",
)


class TemperatureTimeseriesError(ValueError):
    def __init__(self, reason: str) -> None:
        super().__init__(reason)
        self.reason = reason


def parse_temperature_timeseries(content: bytes, filename: str) -> dict[str, object]:
    suffix = Path(filename).suffix.casefold()
    if suffix == ".csv":
        try:
            rows = list(csv.reader(StringIO(content.decode("utf-8-sig"), newline="")))
        except (UnicodeDecodeError, csv.Error) as exc:
            raise TemperatureTimeseriesError("parse") from exc
    elif suffix == ".xlsx":
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            try:
                rows = list(workbook.active.iter_rows(values_only=True))
            finally:
                workbook.close()
        except Exception as exc:
            raise TemperatureTimeseriesError("parse") from exc
    else:
        raise TemperatureTimeseriesError("extension")
    return _rows_metadata(rows)


def ensure_temperature_timeseries_metadata(
    file_asset: FileAsset,
    storage: FileStorageService,
) -> bool:
    metadata = file_asset.metadata_json or {}
    if all(key in metadata for key in TIMESERIES_METADATA_KEYS):
        if not temperature_timeseries_metadata_is_valid(metadata):
            raise TemperatureTimeseriesError("metadata")
        return False
    try:
        content = storage.resolve(file_asset.storage_path).read_bytes()
    except (OSError, ValueError) as exc:
        raise TemperatureTimeseriesError("storage") from exc
    parsed = parse_temperature_timeseries(content, file_asset.original_name)
    file_asset.metadata_json = {**metadata, **parsed}
    return True


def temperature_timeseries_metadata_is_valid(metadata: dict[str, Any]) -> bool:
    columns = metadata.get("columns")
    numeric_columns = metadata.get("numeric_columns")
    numeric_column_pairs = metadata.get("numeric_column_pairs")
    row_count = metadata.get("row_count")
    if (
        not isinstance(columns, list)
        or not columns
        or not all(isinstance(column, str) and column for column in columns)
        or len(columns) != len(set(columns))
        or not isinstance(numeric_columns, list)
        or len(numeric_columns) < 2
        or not all(isinstance(column, str) for column in numeric_columns)
        or len(numeric_columns) != len(set(numeric_columns))
        or not set(numeric_columns).issubset(columns)
        or not isinstance(numeric_column_pairs, list)
        or not numeric_column_pairs
        or not isinstance(row_count, int)
        or isinstance(row_count, bool)
        or row_count < 1
    ):
        return False
    pairs = [
        frozenset(pair)
        for pair in numeric_column_pairs
        if isinstance(pair, list)
        and len(pair) == 2
        and all(isinstance(column, str) for column in pair)
    ]
    return (
        len(pairs) == len(numeric_column_pairs)
        and all(len(pair) == 2 and pair.issubset(numeric_columns) for pair in pairs)
        and len(pairs) == len(set(pairs))
    )


def temperature_timeseries_mapping_error(
    metadata: dict[str, Any],
    reference: dict[str, Any],
) -> str | None:
    if not temperature_timeseries_metadata_is_valid(metadata):
        return "file_metadata"
    numeric_columns = set(metadata["numeric_columns"])
    time_column = reference.get("time_column")
    if time_column not in numeric_columns:
        return "time_column"
    channels = reference.get("channels")
    if not isinstance(channels, list) or not channels:
        return "column_name"
    channel_columns = [
        channel.get("column_name") if isinstance(channel, dict) else None for channel in channels
    ]
    if any(column not in numeric_columns for column in channel_columns):
        return "column_name"
    if time_column in channel_columns:
        return "column_reuse"
    paired_columns = {frozenset(pair) for pair in metadata["numeric_column_pairs"]}
    if any(frozenset((time_column, column)) not in paired_columns for column in channel_columns):
        return "column_pair"
    return None


def _rows_metadata(rows: list[tuple[Any, ...]] | list[list[str]]) -> dict[str, object]:
    while rows and not any(_present(cell) for cell in rows[-1]):
        rows.pop()
    if len(rows) < 2:
        raise TemperatureTimeseriesError("rows")
    width = max(
        (index + 1 for row in rows for index, cell in enumerate(row) if _present(cell)),
        default=0,
    )
    if width == 0:
        raise TemperatureTimeseriesError("header")
    header = [_header(rows[0][index] if index < len(rows[0]) else None) for index in range(width)]
    if any(not column for column in header) or len(header) != len(set(header)):
        raise TemperatureTimeseriesError("header")
    data_rows = [
        [row[index] if index < len(row) else None for index in range(width)]
        for row in rows[1:]
        if any(_present(cell) for cell in row)
    ]
    if not data_rows:
        raise TemperatureTimeseriesError("rows")
    numeric_columns = [
        header[index]
        for index in range(width)
        if (values := [row[index] for row in data_rows if _present(row[index])])
        and all(_numeric(value) for value in values)
    ]
    if len(numeric_columns) < 2:
        raise TemperatureTimeseriesError("numeric_columns")
    numeric_column_pairs = [
        [header[left], header[right]]
        for left in range(width)
        if header[left] in numeric_columns
        for right in range(left + 1, width)
        if header[right] in numeric_columns
        and any(_numeric(row[left]) and _numeric(row[right]) for row in data_rows)
    ]
    if not numeric_column_pairs:
        raise TemperatureTimeseriesError("numeric_column_pairs")
    return {
        "columns": header,
        "numeric_columns": numeric_columns,
        "numeric_column_pairs": numeric_column_pairs,
        "row_count": len(data_rows),
    }


def _present(value: Any) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def _header(value: Any) -> str:
    return str(value).strip() if _present(value) else ""


def _numeric(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, Real):
        return math.isfinite(float(value))
    if isinstance(value, str):
        try:
            return math.isfinite(float(value.strip()))
        except ValueError:
            return False
    return False
