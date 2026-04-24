from __future__ import annotations

import json
from collections.abc import Iterable
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.models.experiment import ExperimentRun
from app.models.module_payload import normalize_module_payload
from app.repositories.file_asset_repository import FileAssetRepository
from app.repositories.module_payload_repository import ModulePayloadRepository
from app.repositories.sample_repository import SampleRepository
from app.schemas.experiment import (
    ExperimentAnalysisCharacterizationRow,
    ExperimentAnalysisExperimentRow,
    ExperimentAnalysisExportRead,
    ExperimentAnalysisFileRow,
    ExperimentAnalysisFurnacePointRow,
    ExperimentAnalysisGasComponentRow,
    ExperimentAnalysisGasProgramRow,
    ExperimentAnalysisGasSegmentRow,
    ExperimentAnalysisPrecursorRow,
    ExperimentAnalysisSampleRow,
    ExperimentAnalysisSubstrateRow,
    ExperimentExportCounts,
    ExperimentExportProvenance,
    ExperimentExportRead,
    ExperimentRead,
)
from app.schemas.module_payload import ExperimentModulePayloadRead
from app.schemas.sample import SampleRead
from app.services.audit_service import AuditService
from app.services.file_asset_service import to_file_asset_read_model


class ExperimentExportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.audit = AuditService(db)
        self.files = FileAssetRepository(db)
        self.module_payloads = ModulePayloadRepository(db)
        self.samples = SampleRepository(db)

    def build_json_export(self, experiment: ExperimentRun) -> ExperimentExportRead:
        modules = [
            ExperimentModulePayloadRead(
                id=item.id,
                experiment_run_id=item.experiment_run_id,
                module_key=item.module_key,
                schema_version=item.schema_version,
                payload_json=normalize_module_payload(item.module_key, item.payload_json),
                note=item.note,
                created_at=item.created_at,
                updated_at=item.updated_at,
            )
            for item in self.module_payloads.list_by_run(experiment.id)
        ]
        samples = [
            SampleRead.model_validate(item)
            for item in self.samples.list_by_experiment(experiment.id, include_deleted=True)
        ]
        files = [
            to_file_asset_read_model(item) for item in self.files.list_by_experiment(experiment.id)
        ]
        audit_refs = [("experiment_run", experiment.id)]
        audit_refs.extend(("sample", sample.id) for sample in samples)
        audit_refs.extend(("file_asset", file_asset.id) for file_asset in files)
        audit_events = self.audit.list_events_for_entities(audit_refs)

        return ExperimentExportRead(
            export_version="cvd_export_v1",
            exported_at=datetime.now(UTC),
            experiment=ExperimentRead.model_validate(experiment),
            modules=modules,
            samples=samples,
            files=files,
            features=[],
            provenance=ExperimentExportProvenance(
                derived_from_run_id=experiment.derived_from_run_id,
                derived_from_run_code=(
                    experiment.derived_from_run.run_code
                    if experiment.derived_from_run is not None
                    else None
                ),
            ),
            audit_events=audit_events,
            counts=ExperimentExportCounts(
                modules=len(modules),
                samples=len(samples),
                files=len(files),
                audit_events=len(audit_events),
            ),
        )

    def build_analysis_export(self, experiment: ExperimentRun) -> ExperimentAnalysisExportRead:
        export_payload = self.build_json_export(experiment)
        payloads = self._module_map(export_payload)
        context = {
            "experiment_id": export_payload.experiment.id,
            "run_code": export_payload.experiment.run_code,
        }

        return ExperimentAnalysisExportRead(
            export_version="cvd_analysis_v1",
            exported_at=datetime.now(UTC),
            experiment=ExperimentAnalysisExperimentRow(
                **context,
                owner_id=export_payload.experiment.owner_id,
                derived_from_run_id=export_payload.experiment.derived_from_run_id,
                derived_from_run_code=export_payload.experiment.derived_from_run_code,
                experiment_type=export_payload.experiment.experiment_type,
                material_system=export_payload.experiment.material_system,
                experiment_date=export_payload.experiment.experiment_date,
                objective=export_payload.experiment.objective,
                status=export_payload.experiment.status,
                quality_label=export_payload.experiment.quality_label,
                summary_result=export_payload.experiment.summary_result,
                invalid_reason=export_payload.experiment.invalid_reason,
                created_at=export_payload.experiment.created_at,
                updated_at=export_payload.experiment.updated_at,
                submitted_at=export_payload.experiment.submitted_at,
                locked_at=export_payload.experiment.locked_at,
            ),
            precursor_rows=self._build_precursor_rows(payloads, context),
            substrate_rows=self._build_substrate_rows(payloads, context),
            furnace_point_rows=self._build_furnace_point_rows(payloads, context),
            gas_program_rows=self._build_gas_program_rows(payloads, context),
            gas_segment_rows=self._build_gas_segment_rows(payloads, context),
            gas_component_rows=self._build_gas_component_rows(payloads, context),
            characterization_rows=self._build_characterization_rows(payloads, context),
            sample_rows=self._build_sample_rows(export_payload, context),
            file_rows=self._build_file_rows(export_payload, context),
        )

    def build_excel_bytes(self, export_payload: ExperimentExportRead) -> bytes:
        workbook = Workbook()
        workbook.remove(workbook.active)

        self._write_basic_info_sheet(workbook.create_sheet("Basic Info"), export_payload)
        self._write_environment_sheet(
            workbook.create_sheet("Environment & Precheck"), export_payload
        )
        self._write_precursors_sheet(workbook.create_sheet("Precursors"), export_payload)
        self._write_substrates_sheet(workbook.create_sheet("Substrates"), export_payload)
        self._write_furnace_sheet(workbook.create_sheet("Furnace Program"), export_payload)
        self._write_gas_sheet(workbook.create_sheet("Gas Program"), export_payload)
        self._write_characterization_sheet(
            workbook.create_sheet("Characterization"), export_payload
        )
        self._write_files_sheet(workbook.create_sheet("Files"), export_payload)
        self._write_audit_sheet(workbook.create_sheet("Audit"), export_payload)

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _write_basic_info_sheet(
        self, worksheet: Worksheet, export_payload: ExperimentExportRead
    ) -> None:
        worksheet.append(["Field", "Value"])
        experiment = export_payload.experiment.model_dump(mode="json")
        ordered_fields = [
            "run_code",
            "id",
            "owner_id",
            "derived_from_run_id",
            "experiment_type",
            "material_system",
            "experiment_date",
            "objective",
            "status",
            "quality_label",
            "summary_result",
            "invalid_reason",
            "created_at",
            "updated_at",
            "submitted_at",
            "locked_at",
        ]
        for key in ordered_fields:
            value = experiment.get(key)
            worksheet.append([key, self._serialize_cell(value)])

    def _write_environment_sheet(
        self, worksheet: Worksheet, export_payload: ExperimentExportRead
    ) -> None:
        worksheet.append(["Module", "Field", "Value"])
        payloads = self._module_map(export_payload)
        for module_key in ("environment", "precheck", "process_observation"):
            for field, value in self._flatten_mapping(payloads.get(module_key, {})).items():
                worksheet.append([module_key, field, self._serialize_cell(value)])

    def _write_precursors_sheet(
        self, worksheet: Worksheet, export_payload: ExperimentExportRead
    ) -> None:
        payloads = self._module_map(export_payload)
        self._write_list_of_dicts(
            worksheet,
            payloads.get("precursors", {}).get("items", []),
        )

    def _write_substrates_sheet(
        self, worksheet: Worksheet, export_payload: ExperimentExportRead
    ) -> None:
        payloads = self._module_map(export_payload)
        self._write_list_of_dicts(
            worksheet,
            payloads.get("substrates", {}).get("items", []),
        )

    def _write_furnace_sheet(
        self, worksheet: Worksheet, export_payload: ExperimentExportRead
    ) -> None:
        payloads = self._module_map(export_payload)
        rows: list[dict[str, Any]] = []
        for zone in payloads.get("furnace_program", {}).get("zones", []):
            base = {
                "zone_index": zone.get("zone_index"),
                "precursor_placed": zone.get("precursor_placed"),
                "zone_note": zone.get("note"),
            }
            for point in zone.get("temperature_program", []):
                rows.append(
                    {
                        **base,
                        "time_min": point.get("time_min"),
                        "temperature_C": point.get("temperature_C"),
                    }
                )
        self._write_list_of_dicts(worksheet, rows)

    def _write_gas_sheet(self, worksheet: Worksheet, export_payload: ExperimentExportRead) -> None:
        payloads = self._module_map(export_payload)
        rows: list[dict[str, Any]] = []
        gas_program = payloads.get("gas_program", {})
        for segment in gas_program.get("segments", []):
            rows.append(
                {
                    "pre_washing_gas": gas_program.get("pre_washing_gas"),
                    "stage": segment.get("stage"),
                    "start_min": segment.get("start_min"),
                    "end_min": segment.get("end_min"),
                    "gas": segment.get("gas"),
                    "flow_sccm": segment.get("flow_sccm"),
                    "components": segment.get("components"),
                }
            )
        self._write_list_of_dicts(worksheet, rows)

    def _write_characterization_sheet(
        self,
        worksheet: Worksheet,
        export_payload: ExperimentExportRead,
    ) -> None:
        payloads = self._module_map(export_payload)
        characterization = payloads.get("characterization", {})
        methods = characterization.get("methods")
        if isinstance(methods, list):
            self._write_list_of_dicts(worksheet, methods)
            return

        worksheet.append(["Field", "Value"])
        for field, value in self._flatten_mapping(characterization).items():
            worksheet.append([field, self._serialize_cell(value)])

    def _write_files_sheet(
        self, worksheet: Worksheet, export_payload: ExperimentExportRead
    ) -> None:
        worksheet.append(
            [
                "original_name",
                "method",
                "file_category",
                "sample_id",
                "content_type",
                "size_bytes",
                "sha256",
                "note",
            ]
        )
        for file_asset in export_payload.files:
            worksheet.append(
                [
                    file_asset.original_name,
                    file_asset.method,
                    file_asset.file_category,
                    str(file_asset.sample_id) if file_asset.sample_id else None,
                    file_asset.content_type,
                    file_asset.size_bytes,
                    file_asset.sha256,
                    file_asset.note,
                ]
            )

    def _write_audit_sheet(
        self, worksheet: Worksheet, export_payload: ExperimentExportRead
    ) -> None:
        worksheet.append(["created_at", "action", "actor_id", "reason"])
        for event in export_payload.audit_events:
            worksheet.append(
                [
                    event.created_at.isoformat(),
                    event.action,
                    str(event.actor_id),
                    event.reason,
                ]
            )

    def _write_list_of_dicts(self, worksheet: Worksheet, rows: Iterable[dict[str, Any]]) -> None:
        normalized_rows = [self._flatten_mapping(row) for row in rows if isinstance(row, dict)]
        if not normalized_rows:
            worksheet.append(["empty"])
            return

        headers = sorted({key for row in normalized_rows for key in row})
        worksheet.append(headers)
        for row in normalized_rows:
            worksheet.append([self._serialize_cell(row.get(header)) for header in headers])

    def _build_precursor_rows(
        self,
        payloads: dict[str, dict[str, Any]],
        context: dict[str, Any],
    ) -> list[ExperimentAnalysisPrecursorRow]:
        items = self._list_payload_items(payloads.get("precursors", {}).get("items"))
        return [
            ExperimentAnalysisPrecursorRow(
                **context,
                precursor_index=index,
                role=item.get("role"),
                type=item.get("type"),
                brand=item.get("brand"),
                concentration=item.get("concentration"),
                concentration_unit=item.get("concentration_unit"),
                method=item.get("method"),
                melting_temperature_C=item.get("melting_temperature_C"),
                spin_speed_rpm=item.get("spin_speed_rpm"),
                pre_spin_speed_rpm=item.get("pre_spin_speed_rpm"),
                preparation_time_min=item.get("preparation_time_min"),
                mass_mg=item.get("mass_mg"),
                batch_no=item.get("batch_no"),
            )
            for index, item in enumerate(items)
        ]

    def _build_substrate_rows(
        self,
        payloads: dict[str, dict[str, Any]],
        context: dict[str, Any],
    ) -> list[ExperimentAnalysisSubstrateRow]:
        items = self._list_payload_items(payloads.get("substrates", {}).get("items"))
        rows: list[ExperimentAnalysisSubstrateRow] = []
        for index, item in enumerate(items):
            treatment_params = item.get("treatment_params")
            if not isinstance(treatment_params, dict):
                treatment_params = {}
            rows.append(
                ExperimentAnalysisSubstrateRow(
                    **context,
                    substrate_index=index,
                    role=item.get("role"),
                    type=item.get("type"),
                    brand=item.get("brand"),
                    size_mm=item.get("size_mm"),
                    treatment_method=item.get("treatment_method"),
                    position_mm=item.get("position_mm"),
                    treatment_params_temperature_C=treatment_params.get("temperature_C"),
                    treatment_params_duration_min=treatment_params.get("duration_min"),
                    treatment_params_power_W=treatment_params.get("power_W"),
                    treatment_params_gas=treatment_params.get("gas"),
                )
            )
        return rows

    def _build_furnace_point_rows(
        self,
        payloads: dict[str, dict[str, Any]],
        context: dict[str, Any],
    ) -> list[ExperimentAnalysisFurnacePointRow]:
        zones = self._list_payload_items(payloads.get("furnace_program", {}).get("zones"))
        rows: list[ExperimentAnalysisFurnacePointRow] = []
        for zone_payload_index, zone in enumerate(zones):
            temperature_program = self._list_payload_items(zone.get("temperature_program"))
            for point_index, point in enumerate(temperature_program):
                rows.append(
                    ExperimentAnalysisFurnacePointRow(
                        **context,
                        furnace_zone_index=zone_payload_index,
                        zone_index=zone.get("zone_index"),
                        temperature_point_index=point_index,
                        precursor_placed=zone.get("precursor_placed"),
                        zone_note=zone.get("note"),
                        time_min=point.get("time_min"),
                        temperature_C=point.get("temperature_C"),
                    )
                )
        return rows

    def _build_gas_program_rows(
        self,
        payloads: dict[str, dict[str, Any]],
        context: dict[str, Any],
    ) -> list[ExperimentAnalysisGasProgramRow]:
        gas_program = payloads.get("gas_program", {})
        if "pre_washing_gas" not in gas_program:
            return []
        return [
            ExperimentAnalysisGasProgramRow(
                **context,
                gas_program_index=0,
                pre_washing_gas=gas_program.get("pre_washing_gas"),
            )
        ]

    def _build_gas_segment_rows(
        self,
        payloads: dict[str, dict[str, Any]],
        context: dict[str, Any],
    ) -> list[ExperimentAnalysisGasSegmentRow]:
        gas_program = payloads.get("gas_program", {})
        segments = self._list_payload_items(gas_program.get("segments"))
        rows: list[ExperimentAnalysisGasSegmentRow] = []
        for index, segment in enumerate(segments):
            rows.append(
                ExperimentAnalysisGasSegmentRow(
                    **context,
                    gas_segment_index=index,
                    pre_washing_gas=gas_program.get("pre_washing_gas"),
                    stage=segment.get("stage"),
                    start_min=segment.get("start_min"),
                    end_min=segment.get("end_min"),
                    gas=segment.get("gas"),
                    flow_sccm=segment.get("flow_sccm"),
                    note=segment.get("note"),
                    component_count=len(self._list_payload_items(segment.get("components"))),
                )
            )
        return rows

    def _build_gas_component_rows(
        self,
        payloads: dict[str, dict[str, Any]],
        context: dict[str, Any],
    ) -> list[ExperimentAnalysisGasComponentRow]:
        segments = self._list_payload_items(payloads.get("gas_program", {}).get("segments"))
        rows: list[ExperimentAnalysisGasComponentRow] = []
        for segment_index, segment in enumerate(segments):
            components = self._list_payload_items(segment.get("components"))
            for component_index, component in enumerate(components):
                rows.append(
                    ExperimentAnalysisGasComponentRow(
                        **context,
                        gas_segment_index=segment_index,
                        gas_component_index=component_index,
                        stage=segment.get("stage"),
                        segment_gas=segment.get("gas"),
                        component_name=component.get("name"),
                        component_gas=component.get("gas"),
                        fraction=component.get("fraction"),
                        ratio_percent=component.get("ratio_percent"),
                    )
                )
        return rows

    def _build_characterization_rows(
        self,
        payloads: dict[str, dict[str, Any]],
        context: dict[str, Any],
    ) -> list[ExperimentAnalysisCharacterizationRow]:
        methods = self._list_payload_items(payloads.get("characterization", {}).get("methods"))
        return [
            ExperimentAnalysisCharacterizationRow(
                **context,
                characterization_index=index,
                method=item.get("method"),
                result=item.get("result"),
                enabled=item.get("enabled"),
                excitation_nm=item.get("excitation_nm"),
                note=item.get("note"),
            )
            for index, item in enumerate(methods)
        ]

    def _build_sample_rows(
        self,
        export_payload: ExperimentExportRead,
        context: dict[str, Any],
    ) -> list[ExperimentAnalysisSampleRow]:
        return [
            ExperimentAnalysisSampleRow(
                **context,
                sample_id=sample.id,
                sample_code=sample.sample_code,
                parent_sample_id=sample.parent_sample_id,
                role=sample.role,
                substrate_type=sample.substrate_type,
                brand=sample.brand,
                size_mm=sample.size_mm,
                treatment=sample.treatment,
                position_mm=sample.position_mm,
                storage_location=sample.storage_location,
                metadata_json_text=self._json_text(sample.metadata_json),
                created_at=sample.created_at,
                updated_at=sample.updated_at,
                deleted_at=sample.deleted_at,
                deleted_by_id=sample.deleted_by_id,
                is_deleted=sample.is_deleted,
            )
            for sample in export_payload.samples
        ]

    def _build_file_rows(
        self,
        export_payload: ExperimentExportRead,
        context: dict[str, Any],
    ) -> list[ExperimentAnalysisFileRow]:
        return [
            ExperimentAnalysisFileRow(
                **context,
                file_id=file_asset.id,
                sample_id=file_asset.sample_id,
                original_name=file_asset.original_name,
                method=file_asset.method,
                file_category=file_asset.file_category,
                content_type=file_asset.content_type,
                size_bytes=file_asset.size_bytes,
                sha256=file_asset.sha256,
                note=file_asset.note,
                metadata_json_text=self._json_text(file_asset.metadata_json),
                created_at=file_asset.created_at,
                updated_at=file_asset.updated_at,
            )
            for file_asset in export_payload.files
        ]

    def _list_payload_items(self, value: Any) -> list[dict[str, Any]]:
        if not isinstance(value, list):
            return []
        return [item for item in value if isinstance(item, dict)]

    def _json_text(self, value: dict[str, Any]) -> str:
        return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))

    def _module_map(self, export_payload: ExperimentExportRead) -> dict[str, dict[str, Any]]:
        return {module.module_key: module.payload_json for module in export_payload.modules}

    def _flatten_mapping(self, mapping: dict[str, Any], prefix: str = "") -> dict[str, Any]:
        flattened: dict[str, Any] = {}
        for key, value in mapping.items():
            composite_key = f"{prefix}.{key}" if prefix else str(key)
            if isinstance(value, dict):
                flattened.update(self._flatten_mapping(value, composite_key))
            else:
                flattened[composite_key] = value
        return flattened

    def _serialize_cell(self, value: Any) -> Any:
        if isinstance(value, list | dict):
            return json.dumps(value, ensure_ascii=False)
        return value
