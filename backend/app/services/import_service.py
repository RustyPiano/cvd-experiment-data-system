from io import BytesIO

import openpyxl
from fastapi import HTTPException, status
from pydantic import ValidationError
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models.module_payload import ExperimentModuleKey, normalize_module_payload
from app.models.user import User, UserRole
from app.schemas.imports import (
    ImportCommitRequest,
    ImportCommitResponse,
    ImportCommitResultItem,
    ImportPreviewResponse,
    ImportProfileListResponse,
)
from app.schemas.module_payload import validate_module_payload
from app.services.experiment_service import ExperimentService
from app.services.imports.registry import get_import_profile, list_import_profiles


class ImportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.experiments = ExperimentService(db)

    def list_profiles(self) -> ImportProfileListResponse:
        return ImportProfileListResponse(
            profiles=[profile.info() for profile in list_import_profiles()]
        )

    def preview(self, *, content: bytes, profile_key: str) -> ImportPreviewResponse:
        max_bytes = get_settings().file_upload_max_bytes
        if len(content) > max_bytes:
            raise HTTPException(
                status_code=status.HTTP_413_CONTENT_TOO_LARGE,
                detail=f"Uploaded file exceeds {max_bytes} bytes",
            )
        profile = self._require_profile(profile_key)
        workbook = self._load_workbook(content)
        try:
            drafts, global_warnings = profile.parse(workbook)
        finally:
            workbook.close()
        return ImportPreviewResponse(
            profile_key=profile_key,
            drafts=drafts,
            global_warnings=global_warnings,
        )

    def commit(
        self,
        payload: ImportCommitRequest,
        current_user: User,
    ) -> ImportCommitResponse:
        if current_user.role == UserRole.VIEWER:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Insufficient permissions",
            )
        self._require_profile(payload.profile_key)

        # Validate every module payload up front and reject the entire request
        # before persisting anything if any row is malformed. This catches the
        # common failure mode (bad data) without committing partial results.
        # Persistence below commits per row (see ``create_from_import``), so a
        # rare failure during write-out — e.g. run-code allocation exhausting
        # its retries — leaves earlier rows committed; the response only lists
        # the rows that were successfully created.
        validation_errors: list[dict[str, object]] = []
        for draft in payload.drafts:
            for module_key_value, module_payload in draft.module_payloads.items():
                if module_key_value == ExperimentModuleKey.BASIC_INFO.value:
                    continue
                if module_key_value not in {key.value for key in ExperimentModuleKey}:
                    continue
                if not isinstance(module_payload, dict):
                    continue
                try:
                    normalized = normalize_module_payload(module_key_value, module_payload)
                    validate_module_payload(module_key_value, normalized)
                except ValidationError as exc:
                    validation_errors.append(
                        {
                            "source_row": draft.source_row,
                            "module_key": module_key_value,
                            "errors": exc.errors(),
                        }
                    )
        if validation_errors:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                detail=validation_errors,
            )

        created: list[ImportCommitResultItem] = []
        for draft in payload.drafts:
            experiment = self.experiments.create_from_import(
                run_level=draft.run_level,
                module_payloads=draft.module_payloads,
                current_user=current_user,
            )
            created.append(
                ImportCommitResultItem(
                    source_row=draft.source_row,
                    experiment_id=experiment.id,
                    run_code=experiment.run_code,
                )
            )
        return ImportCommitResponse(created=created)

    def _require_profile(self, profile_key: str):
        profile = get_import_profile(profile_key)
        if profile is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Unknown import profile: {profile_key}",
            )
        return profile

    def _load_workbook(self, content: bytes):
        try:
            return openpyxl.load_workbook(BytesIO(content), data_only=True)
        except Exception as exc:  # noqa: BLE001 - surface a friendly parse error
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                detail="无法读取 Excel 文件，请确认文件为有效的 .xlsx 格式。",
            ) from exc
